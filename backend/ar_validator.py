"""
AR (Accounts Receivable) migration validator.

Compares an ECC AR registry against a filled-in S/4 "Customer Open Items"
template and runs a fixed set of reconciliation checks between them.

Design notes
------------
Every check returns a dict of the shape::

    {
        "check_name": str,
        "status": "PASS" | "FAIL",
        "message": str,
        "details": [ <detail>, ... ],   # may be empty
        ... check-specific extra fields ...
    }

and every entry in ``details`` shares one schema::

    {
        "label": str,
        "left_count": number | None,   # ECC side
        "right_count": number | None,  # S/4 side
        "status": "PASS" | "FAIL" | "MAPPING_ERROR",
        "message": str,
        "money": bool,                 # optional, present only when True
        ... detail-specific extra fields ...
    }

This lets a single frontend component render every check the same way
(see ValidationScoreboard.jsx), and lets ``validate_ar_files`` drive every
check through the same loop instead of hand-rolled per-check wiring.

Adding a 6th check means writing one ``validate_*`` function with this
shape and adding it to ``CHECK_FUNCTIONS`` below -- nothing else changes.
"""

from typing import Any, Callable, Dict, Iterable, List, Optional

import pandas as pd
import openpyxl

import mappings


# ============================================================
# AR Validation Configuration
# ============================================================

ECC_START_ROW = 2
S4_START_ROW = 9

S4_SHEET_NAME = "Customer Open Items"

COMPANY_CODE_MAPPING = mappings.COMPANY_CODE_MAPPING


# ============================================================
# Payment Terms Group Configuration
# ============================================================
# Each set contains approximately 5 S/4 payment terms. If
# multiple ECC payment terms map to the same S/4 term, all of
# those ECC terms remain in the same set.
# ============================================================

PAYMENT_TERM_GROUPS = [
    {"set": 1, "ecc": {"001", "003", "004", "014", "015"}, "s4": {"P210", "Z200", "P215", "P220", "Z251"}},
    {"set": 2, "ecc": {"016", "017", "018", "019", "020"}, "s4": {"Z291", "Z261", "Z245", "Z230", "Z231"}},
    {"set": 3, "ecc": {"021", "022", "023", "024", "025"}, "s4": {"Z232", "Z246", "Z233", "Z260", "Z305"}},
    {"set": 4, "ecc": {"026", "027", "029", "060", "030", "035"}, "s4": {"NT12", "Z160", "Z262", "Z276", "Z290"}},
    {"set": 5, "ecc": {"036", "038", "039", "040", "444", "041", "048", "401"}, "s4": {"Z130", "P030", "P230", "NT30", "NT60"}},
    {"set": 6, "ecc": {"042", "442", "043", "443", "044", "445", "045", "046"}, "s4": {"NT90", "NT45", "NT75", "NT15", "P025"}},
    {"set": 7, "ecc": {"050", "052", "056", "058", "059"}, "s4": {"Z400", "Z132", "Z216", "Z265", "Z263"}},
    {"set": 8, "ecc": {"061", "062", "063", "064", "065"}, "s4": {"Z264", "Z247", "Z161", "Z330", "Z146"}},
    {"set": 9, "ecc": {"070", "072", "073", "075", "091"}, "s4": {"P260", "P225", "Z163", "Z505", "NTLC"}},
    {"set": 10, "ecc": {"094", "097", "100", "107", "109"}, "s4": {"Z346", "Z164", "Z167", "Z316", "Z225"}},
    {"set": 11, "ecc": {"111", "112", "114", "115", "117"}, "s4": {"Z234", "Z235", "P190", "P160", "P101"}},
    {"set": 12, "ecc": {"118", "119", "122", "33", "129", "402", "138"}, "s4": {"Z162", "Z212", "Z131", "NT10", "Z165"}},
    {"set": 13, "ecc": {"139", "141", "400", "403", "441"}, "s4": {"Z176", "E225", "Z166", "NT00", "NT65"}},
]


# ============================================================
# Utility Functions
# ============================================================

def clean_string(value: Any) -> str:
    """
    Convert a cell value to a clean string.

    Empty/NaN values become "".
    Numeric values such as 1000.0 become "1000".
    """
    if value is None or pd.isna(value):
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def is_non_empty(value: Any) -> bool:
    """
    Returns True when an Excel cell contains an actual value.
    """
    if value is None:
        return False

    if isinstance(value, str):
        return value.strip() != ""

    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass

    return True


def normalize_series(series: pd.Series, upper: bool = True) -> pd.Series:
    """
    Apply clean_string element-wise, optionally upper-casing the result.

    Centralizes the `.apply(clean_string).str.upper()` pattern that
    every count-comparison check needs before comparing ECC and S/4
    values, so each check just declares which columns it cares about.
    """
    cleaned = series.apply(clean_string)
    return cleaned.str.upper() if upper else cleaned


# ============================================================
# Result-Building Helpers
# ============================================================
# Shared by every check below so each one only has to express its own
# comparison logic, not the shape of the dict it returns.
# ============================================================

def pass_fail(is_match: bool) -> str:
    """Map a boolean match into the "PASS" / "FAIL" status string."""
    return "PASS" if is_match else "FAIL"


def make_detail(
    label: str,
    left_count: Optional[float],
    right_count: Optional[float],
    message: str,
    status: Optional[str] = None,
    money: bool = False,
    **extra: Any,
) -> Dict[str, Any]:
    """
    Build one row of a check's `details` list.

    `status` defaults to a straight left == right comparison; pass it
    explicitly for cases that need other logic (e.g. a missing mapping).
    """
    if status is None:
        status = pass_fail(left_count == right_count)

    detail: Dict[str, Any] = {
        "label": label,
        "left_count": left_count,
        "right_count": right_count,
        "status": status,
        "message": message,
    }

    if money:
        detail["money"] = True

    detail.update(extra)
    return detail


def make_check(
    check_name: str,
    details: List[Dict[str, Any]],
    passing_message: str,
    failing_message: str,
    **extra: Any,
) -> Dict[str, Any]:
    """
    Build a check's top-level result from its already-built `details`.

    Overall status is PASS only when every detail row is PASS. Callers
    that need a dynamic failing_message (e.g. naming which sets failed)
    build that string themselves before calling this.
    """
    overall_pass = all(detail["status"] == "PASS" for detail in details)

    check: Dict[str, Any] = {
        "check_name": check_name,
        "status": pass_fail(overall_pass),
        "details": details,
        "message": passing_message if overall_pass else failing_message,
    }
    check.update(extra)
    return check


def missing_column_check(check_name: str, source_label: str, column: str) -> Dict[str, Any]:
    """Short-circuit result for a check whose required column is absent."""
    return {
        "check_name": check_name,
        "status": "FAIL",
        "details": [],
        "message": f'{source_label} does not contain a "{column}" column.',
    }


def first_missing_column(df: pd.DataFrame, columns: Iterable[str]) -> Optional[str]:
    """Return the first of `columns` not present in `df`, or None."""
    for column in columns:
        if column not in df.columns:
            return column
    return None


# ============================================================
# Excel Reading
# ============================================================

def read_ecc_registry(registry_file) -> pd.DataFrame:
    """
    Read the ECC AR registry.

    ECC data begins from Excel row 2, so row 2 is treated
    as the header row and records begin from row 3.
    """

    df = pd.read_excel(
        registry_file,
        header=0,
    )

    # Remove completely empty rows.
    df = df.dropna(how="all")

    return df


def read_s4_customer_open_items(filled_file) -> pd.DataFrame:
    """
    Read the S/4 Customer Open Items sheet.

    S/4 template structure:
        Row 5 = technical field names (BUKRS, KUNNR, BLART, etc.)
        Row 8 = human-readable field descriptions
        Row 9 onward = actual records

    The validator uses the technical field names so that
    validations can consistently refer to fields such as BUKRS.
    """

    wb = openpyxl.load_workbook(
        filled_file,
        read_only=True,
        data_only=True,
    )

    if S4_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f'S/4 file does not contain the required sheet '
            f'"{S4_SHEET_NAME}".'
        )

    ws = wb[S4_SHEET_NAME]

    # ------------------------------------------------------------
    # Row 5 contains the technical field names:
    # BUKRS, KUNNR, BLART, BLDAT, etc.
    # ------------------------------------------------------------
    technical_headers = [
        clean_string(cell.value)
        for cell in ws[5]
    ]

    # ------------------------------------------------------------
    # Row 8 contains human-readable labels.
    # We don't use these as DataFrame column names, but we
    # retain the structure of the template.
    # ------------------------------------------------------------
    descriptive_headers = [
        clean_string(cell.value)
        for cell in ws[8]
    ]

    # ------------------------------------------------------------
    # Read actual data beginning from row 9.
    # ------------------------------------------------------------
    records = []

    for row in ws.iter_rows(
        min_row=S4_START_ROW,
        values_only=True,
    ):
        # Ignore completely empty rows.
        if not any(is_non_empty(value) for value in row):
            continue

        record = {}

        for index, value in enumerate(row):
            if index < len(technical_headers):
                technical_field = technical_headers[index]

                if technical_field:
                    record[technical_field] = value

        records.append(record)

    wb.close()

    return pd.DataFrame(records)


# ============================================================
# Validation 1
# Total Record Count
# ============================================================

def validate_record_count(ecc_df: pd.DataFrame, s4_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Compare the total number of ECC records against
    the total number of S/4 records.
    """

    ecc_count = len(ecc_df)
    s4_count = len(s4_df)
    difference = ecc_count - s4_count
    status = pass_fail(difference == 0)

    message = (
        f"Record count matches. "
        f"Both ECC and S/4 contain {ecc_count} records."
        if status == "PASS"
        else (
            f"Record count mismatch. "
            f"ECC contains {ecc_count} records while "
            f"S/4 contains {s4_count} records."
        )
    )

    detail = make_detail(
        label="Total records",
        left_count=ecc_count,
        right_count=s4_count,
        status=status,
        message=message,
    )

    return make_check(
        check_name="Total Record Count",
        details=[detail],
        passing_message=message,
        failing_message=message,
        ecc_count=ecc_count,
        s4_count=s4_count,
        difference=difference,
    )


# ============================================================
# Validation 2
# Company Code Distribution
# ============================================================

def validate_company_code_counts(ecc_df: pd.DataFrame, s4_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Dynamically determine all ECC company codes and compare
    their counts against the corresponding S/4 company codes.

    Example:

        ECC       S/4
        US01  ->  1000
        US06  ->  1001
        CA01  ->  1200

    The company codes are NOT hardcoded into the validation
    logic. They are discovered from the registry.
    """

    missing = first_missing_column(ecc_df, ["Company Code"])
    if missing:
        return missing_column_check("Company Code Distribution", "ECC registry", missing)

    missing = first_missing_column(s4_df, ["BUKRS"])
    if missing:
        return missing_column_check("Company Code Distribution", "S/4 file", missing)

    ecc_codes = normalize_series(ecc_df["Company Code"])
    s4_codes = normalize_series(s4_df["BUKRS"])

    # Discover company codes dynamically from ECC.
    distinct_ecc_codes = ecc_codes.replace("", pd.NA).dropna().unique()

    details = []

    for ecc_code in distinct_ecc_codes:
        ecc_code = ecc_code.upper()
        ecc_count = int((ecc_codes == ecc_code).sum())

        # Look up ECC -> S/4 mapping.
        s4_code = COMPANY_CODE_MAPPING.get(ecc_code)

        if not s4_code:
            details.append(make_detail(
                label=f"{ecc_code} (no mapping)",
                left_count=ecc_count,
                right_count=None,
                status="MAPPING_ERROR",
                message=f"No ECC -> S/4 company code mapping exists for {ecc_code}.",
                ecc_code=ecc_code,
                s4_code=None,
            ))
            continue

        s4_code = clean_string(s4_code)
        s4_count = int((s4_codes == s4_code).sum())

        details.append(make_detail(
            label=f"{ecc_code} \u2192 {s4_code}",
            left_count=ecc_count,
            right_count=s4_count,
            message=(
                "Company code count matches."
                if ecc_count == s4_count
                else f"Company code count mismatch: ECC={ecc_count}, S/4={s4_count}."
            ),
            ecc_code=ecc_code,
            s4_code=s4_code,
        ))

    return make_check(
        check_name="Company Code Distribution",
        details=details,
        passing_message="All company code counts match.",
        failing_message="One or more company code counts do not match.",
    )


# ============================================================
# Validation 3
# Sign Validation
# ============================================================

def validate_sign(ecc_df: pd.DataFrame, s4_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Validate the sign of amounts between ECC and S/4.

    ECC:
        Debit/Credit Ind. = S -> amount is positive
        Debit/Credit Ind. = H -> amount is negative

    S/4:
        Positive WRBTR values must match ECC S total.
        Negative WRBTR values must match ECC H total
        in absolute value.
    """

    # --------------------------------------------------------
    # ECC amounts
    # --------------------------------------------------------

    ecc_amounts = pd.to_numeric(ecc_df["Amount"], errors="coerce").fillna(0)
    ecc_indicator = normalize_series(ecc_df["Debit/Credit Ind."])

    ecc_s_total = float(ecc_amounts[ecc_indicator == "S"].abs().sum())
    ecc_h_total = float(ecc_amounts[ecc_indicator == "H"].abs().sum())

    # --------------------------------------------------------
    # S/4 WRBTR amounts
    # --------------------------------------------------------

    s4_amounts = pd.to_numeric(s4_df["WRBTR"], errors="coerce").fillna(0)

    s4_positive_total = float(s4_amounts[s4_amounts > 0].sum())
    s4_negative_total = float(s4_amounts[s4_amounts < 0].abs().sum())

    # --------------------------------------------------------
    # Compare (a small epsilon absorbs float rounding, not a real mismatch)
    # --------------------------------------------------------

    TOLERANCE = 0.01
    s_match = abs(ecc_s_total - s4_positive_total) < TOLERANCE
    h_match = abs(ecc_h_total - s4_negative_total) < TOLERANCE

    details = [
        make_detail(
            label="S / Debit (positive)",
            left_count=ecc_s_total,
            right_count=s4_positive_total,
            status=pass_fail(s_match),
            message=(
                "S (debit) total matches S/4 positive WRBTR total."
                if s_match
                else f"S (debit) mismatch: ECC={ecc_s_total:.2f}, S/4 positive={s4_positive_total:.2f}."
            ),
            money=True,
        ),
        make_detail(
            label="H / Credit (negative)",
            left_count=ecc_h_total,
            right_count=s4_negative_total,
            status=pass_fail(h_match),
            message=(
                "H (credit) total matches S/4 negative WRBTR total."
                if h_match
                else f"H (credit) mismatch: ECC={ecc_h_total:.2f}, S/4 negative={s4_negative_total:.2f}."
            ),
            money=True,
        ),
    ]

    return make_check(
        check_name="Amount Sign Validation",
        details=details,
        passing_message=(
            "ECC S/H amount totals match the corresponding "
            "positive/negative S/4 WRBTR totals."
        ),
        failing_message=(
            "Sign validation failed. One or both ECC S/H totals "
            "do not match the corresponding S/4 positive/negative totals."
        ),
        ecc_s_total=ecc_s_total,
        s4_positive_total=s4_positive_total,
        s_difference=ecc_s_total - s4_positive_total,
        ecc_h_total=ecc_h_total,
        s4_negative_total=-s4_negative_total,
        h_difference=-ecc_h_total + s4_negative_total,
        s_status=details[0]["status"],
        h_status=details[1]["status"],
    )


# ============================================================
# Validation 4
# Payment Terms Blank Count
# ============================================================

def validate_payment_terms_blank_count(ecc_df: pd.DataFrame, s4_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Validate that the number of blank Terms of Payment values
    in the ECC registry matches the number of
    'No payment terms in ECC' values in S/4 ZTERM.
    """

    ecc_blank_count = int(
        ecc_df["Terms of Payment"].apply(lambda value: not is_non_empty(value)).sum()
    )

    s4_no_payment_terms_count = int(
        s4_df["ZTERM"].apply(lambda value: clean_string(value) == "No payment terms in ECC").sum()
    )

    status = pass_fail(ecc_blank_count == s4_no_payment_terms_count)

    message = (
        "Payment terms blank count matches. "
        f"ECC contains {ecc_blank_count} blank Terms of Payment values "
        f"and S/4 contains {s4_no_payment_terms_count} "
        "'No payment terms in ECC' values in ZTERM."
        if status == "PASS"
        else (
            "Payment terms blank count mismatch. "
            f"ECC contains {ecc_blank_count} blank Terms of Payment values "
            f"while S/4 contains {s4_no_payment_terms_count} "
            "'No payment terms in ECC' values in ZTERM."
        )
    )

    detail = make_detail(
        label="Blank / no payment terms",
        left_count=ecc_blank_count,
        right_count=s4_no_payment_terms_count,
        status=status,
        message=message,
    )

    return make_check(
        check_name="Payment Terms Blank Count",
        details=[detail],
        passing_message=message,
        failing_message=message,
        ecc_blank_count=ecc_blank_count,
        s4_no_payment_terms_count=s4_no_payment_terms_count,
        difference=ecc_blank_count - s4_no_payment_terms_count,
    )


# ============================================================
# Validation 5
# Payment Terms Group Count
# ============================================================

def validate_payment_terms_group_counts(ecc_df: pd.DataFrame, s4_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Compare payment-term counts at the configured group level.

    Each ECC row contributes one count to the set containing its
    Terms of Payment. Each S/4 row contributes one count to the set
    containing its ZTERM. The validation passes only when every set
    has the same ECC and S/4 count.
    """

    missing = first_missing_column(ecc_df, ["Terms of Payment"])
    if missing:
        return missing_column_check("Payment Terms Group Count", "ECC registry", missing)

    missing = first_missing_column(s4_df, ["ZTERM"])
    if missing:
        return missing_column_check("Payment Terms Group Count", "S/4 file", missing)

    ecc_terms_series = normalize_series(ecc_df["Terms of Payment"]).str.lstrip("0")
    s4_terms_series = normalize_series(s4_df["ZTERM"])

    details = []

    for group in PAYMENT_TERM_GROUPS:
        set_number = group["set"]
        ecc_terms = {clean_string(term).upper().lstrip("0") for term in group["ecc"]}
        s4_terms = {clean_string(term).upper() for term in group["s4"]}

        ecc_count = int(ecc_terms_series.isin(ecc_terms).sum())
        s4_count = int(s4_terms_series.isin(s4_terms).sum())

        details.append(make_detail(
            label=f"Set {set_number}",
            left_count=ecc_count,
            right_count=s4_count,
            message=(
                f"Set {set_number} count matches."
                if ecc_count == s4_count
                else f"Set {set_number} count mismatch: ECC={ecc_count}, S/4={s4_count}."
            ),
            ecc_terms=sorted(group["ecc"]),
            s4_terms=sorted(group["s4"]),
        ))

    total_ecc_count = sum(detail["left_count"] for detail in details)
    total_s4_count = sum(detail["right_count"] for detail in details)
    failed_sets = [detail["label"] for detail in details if detail["status"] == "FAIL"]

    return make_check(
        check_name="Payment Terms Group Count",
        details=details,
        passing_message="All payment terms groups have matching ECC and S/4 counts.",
        failing_message=(
            "Payment terms group validation failed. Mismatched sets: "
            + ", ".join(failed_sets) + "."
        ),
        ecc_total_count=total_ecc_count,
        s4_total_count=total_s4_count,
        difference=total_ecc_count - total_s4_count,
    )


# ============================================================
# Check Registry
# ============================================================
# The single place that lists which checks run. Adding a check means
# writing one validate_* function above and appending it here -- nothing
# in validate_ar_files needs to change.
# ============================================================

CHECK_FUNCTIONS: List[Callable[[pd.DataFrame, pd.DataFrame], Dict[str, Any]]] = [
    validate_record_count,
    validate_company_code_counts,
    validate_sign,
    validate_payment_terms_blank_count,
    validate_payment_terms_group_counts,
]


# ============================================================
# Overall Validation
# ============================================================

def calculate_overall_status(checks: List[Dict[str, Any]]) -> str:
    """
    Overall validation passes only when every check passes.
    """
    return pass_fail(all(check.get("status") == "PASS" for check in checks))


def summarize_checks(checks: List[Dict[str, Any]]) -> Dict[str, int]:
    """Roll up per-check statuses into the summary block."""
    passed = sum(1 for check in checks if check["status"] == "PASS")
    return {
        "total_checks": len(checks),
        "passed": passed,
        "failed": len(checks) - passed,
    }


# ============================================================
# Main AR Validation Function
# ============================================================

def validate_ar_files(registry_file, filled_file) -> Dict[str, Any]:
    """
    Run every check in CHECK_FUNCTIONS against the ECC registry and the
    filled S/4 template.

    Returns a structured dictionary that can be used by:
        1. FastAPI
        2. Frontend
        3. Future DOCX report generator
    """

    ecc_df = read_ecc_registry(registry_file)
    s4_df = read_s4_customer_open_items(filled_file)

    checks = [check_fn(ecc_df, s4_df) for check_fn in CHECK_FUNCTIONS]

    return {
        "process": "AR",
        "overall_status": calculate_overall_status(checks),
        "summary": summarize_checks(checks),
        "checks": checks,
    }