# backend/asset_processor.py

import io
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import mappings
from validation_utils import extract_mandatory_fields, is_blank
from openpyxl.styles import PatternFill

# Fill used to flag any row that has a blank mandatory field, so it's
# visually obvious in the generated workbook which rows still need
# attention before the file is loaded into S/4.
MISSING_MANDATORY_FILL = PatternFill(
    start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
)


def clean_string(val):
    if pd.isna(val) or val is None:
        return ""

    if isinstance(val, float) and val.is_integer():
        return str(int(val))

    return str(val).strip()


def clean_int(val, default=""):
    if pd.isna(val) or val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return str(val).strip()


def clean_float(val, default=None):
    """
    Convert a value to float, including SAP/Excel "accounting format"
    text such as "5,114.43-" or "(5,114.43)" for negative numbers.
    """
    if pd.isna(val) or val is None:
        return default

    if isinstance(val, (int, float)):
        return float(val)

    text = str(val).strip()
    if not text:
        return default

    negative = False

    if text.endswith('-'):
        negative = True
        text = text[:-1].strip()
    elif text.startswith('(') and text.endswith(')'):
        negative = True
        text = text[1:-1].strip()

    text = text.replace(',', '')

    try:
        num = float(text)
        return -num if negative else num
    except (ValueError, TypeError):
        return val


class RegistryMismatchError(ValueError):
    """Raised when the uploaded file doesn't look like an Asset registry."""
    pass


REQUIRED_ASSET_COLUMNS = ['CoCd', 'Asset', 'Deact.Date', 'Plnt', 'Location']


def validate_asset_registry(df):
    missing = [c for c in REQUIRED_ASSET_COLUMNS if c not in df.columns]
    if missing:
        raise RegistryMismatchError(
            "This file doesn't look like an Asset registry. "
            f"Missing expected column(s): {', '.join(missing)}. "
            "Make sure you selected the right file for the Assets process."
        )


def read_asset_sheets(registry_file):
    """
    Reads every sheet in the workbook and combines the ones that actually
    look like asset registry data (i.e. have all of REQUIRED_ASSET_COLUMNS)
    into a single DataFrame — auto-detected by column shape, not by
    hardcoded sheet names.

    This matters because real registry exports have shown up in more than
    one shape from the same source: one file had a separate sheet per
    company code ("US01", "US06", "CA01"), another has every company code
    combined into a single sheet with its own name ("Asset Registry
    Sheet") and a CoCd column carrying the company code per row. Hardcoding
    either shape's sheet names breaks on the other. Detecting by column
    shape works for both, plus any future variant, without a code change.

    Returns (combined_df, skipped_sheets) where skipped_sheets is a list
    of {"sheet": name, "reason": str} for every tab that was excluded
    (e.g. a cost center lookup tab, a differently-shaped list), so the
    caller can report it instead of silently losing data.
    """
    all_sheets = pd.read_excel(registry_file, sheet_name=None)

    matched_frames = []
    skipped_sheets = []

    for sheet_name, df in all_sheets.items():
        missing = [c for c in REQUIRED_ASSET_COLUMNS if c not in df.columns]
        if missing:
            skipped_sheets.append({
                "sheet": sheet_name,
                "reason": f"Missing expected column(s): {', '.join(missing)}",
            })
            continue

        df = df.copy()
        df['__source_sheet'] = sheet_name
        matched_frames.append(df)

    if not matched_frames:
        raise RegistryMismatchError(
            "This file doesn't look like an Asset registry. None of its "
            f"{len(all_sheets)} sheet(s) contain the expected columns: "
            f"{', '.join(REQUIRED_ASSET_COLUMNS)}. "
            "Make sure you selected the right file for the Assets process."
        )

    combined = pd.concat(matched_frames, ignore_index=True)
    return combined, skipped_sheets


def clean_date(val):
    if pd.isna(val) or val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    val_str = str(val).strip()
    if val_str == "00/00/0000" or val_str == "00.00.0000" or val_str == "" or val_str == "NaT":
        return None

    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%d.%m.%Y', '%Y%m%d'):
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return val


def process_asset_registry(
    registry_file,
    template_path="templates/assets_load_template.xlsx",
    custom_mappings=None,
    sheet_names=None
) -> io.BytesIO:
    """
    Processes the ECC asset registry spreadsheet(s), filters active records,
    maps fields to S/4 HANA, and writes them to the prebuilt Excel template
    sheets starting at Row 9.

    Sheets are auto-detected by column shape (see read_asset_sheets) — the
    `sheet_names` parameter is accepted for backward compatibility but is
    IGNORED; it existed to hardcode per-company-code sheet names, which
    broke the moment a differently-shaped export (single sheet, all
    company codes combined) showed up. Auto-detection handles both.

    Any row whose mandatory fields are blank, OR whose cost center lookup
    failed, gets its entire row filled solid red in every sheet it
    appears on.

    :return: (io.BytesIO, list) – the generated workbook and a list of
        validation errors.
    """
    if custom_mappings is None:
        custom_mappings = {}

    cocd_map = custom_mappings.get('cocd', {})
    plant_loc_map = custom_mappings.get('plant_loc', {})
    cost_center_map = custom_mappings.get('cost_center', {})

    def map_cocd(ecc_cocd):
        ecc_str = clean_string(ecc_cocd)
        if ecc_str in cocd_map:
            return cocd_map[ecc_str]
        return mappings.get_s4_company_code(ecc_cocd)

    def map_plant_loc(ecc_plant, ecc_loc):
        p_val = clean_int(ecc_plant)
        l_str = clean_string(ecc_loc).upper()
        if (p_val, l_str) in plant_loc_map:
            return plant_loc_map[(p_val, l_str)]
        return mappings.get_s4_plant_and_location(ecc_plant, ecc_loc)

    def map_cost_center(ecc_cc, ecc_plant, ecc_loc):
        cc_val = clean_int(ecc_cc)
        if cc_val in cost_center_map:
            return cost_center_map[cc_val]
        return mappings.get_s4_cost_center(ecc_cc, ecc_plant, ecc_loc)

    # 1. Read every sheet, keep only the ones shaped like asset data, and
    # combine them into one DataFrame.
    df_raw, skipped_sheets = read_asset_sheets(registry_file)

    def is_active(row):
        deact_raw = row.get('Deact.Date', '')
        if pd.isna(deact_raw):
            return True
        deact_val = str(deact_raw).strip()
        return deact_val in ('', '00/00/0000', '00.00.0000')

    df_active = df_raw[df_raw.apply(is_active, axis=1)].copy()

    wb = openpyxl.load_workbook(template_path)

    sheets_to_fill = {
        'Master Details': 'S_KEY',
        'Allocations': 'S_ALLOCATIONS',
        'Origin': 'S_ORIGIN',
        'Depreciation Areas': 'S_DEPR',
        'Cumulative Values': 'S_CUM_VAL',
        'Posted Values': 'S_POST_VAL',
        'Posting Information': 'S_POSTINGINFORMATION',
        'Time-Dependent Data': 'S_TIMEDEPENDENTDATA',
        'Acct. Assignmt. for Investment': 'S_INVESTACCTASSIGNMN'
    }

    processed_records = []
    validation_errors = []

    for idx, row in df_active.iterrows():
        ecc_cocd = row.get('CoCd')
        s4_cocd = map_cocd(ecc_cocd)

        ecc_asset = clean_string(row.get('Asset'))
        s4_subnumber = "0"

        ecc_plant = row.get('Plnt')
        ecc_location = row.get('Location')
        pl_mapped = map_plant_loc(ecc_plant, ecc_location)
        s4_plant = pl_mapped['s4_plant']
        s4_location = pl_mapped['s4_location']

        ecc_cc = row.get('Cost Ctr')
        s4_cost_center = map_cost_center(ecc_cc, ecc_plant, ecc_location)
        cost_center_missing = is_blank(s4_cost_center)

        source_row = idx + 2

        record = {
            'ecc_row': row,
            'source_row': source_row,
            'source_sheet': row.get('__source_sheet', ''),
            's4_cocd': s4_cocd,
            's4_asset': ecc_asset,
            's4_subnumber': s4_subnumber,
            's4_plant': s4_plant,
            's4_location': s4_location,
            's4_cost_center': s4_cost_center,
            'cost_center_missing': cost_center_missing,
        }
        processed_records.append(record)

        if cost_center_missing:
            validation_errors.append({
                'sheet': 'Time-Dependent Data',
                'field': 'KOSTL',
                'field_label': 'Cost Center',
                'source_row': source_row,
                'source_sheet': record['source_sheet'],
                'asset': ecc_asset,
                'subnumber': s4_subnumber,
                'company_code': s4_cocd,
            })

    for sheet_name, tech_id in sheets_to_fill.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found in template, skipping.")
            continue

        ws = wb[sheet_name]

        tech_cols = [clean_string(ws.cell(row=5, column=col).value) for col in range(1, ws.max_column + 1)]
        col_to_idx = {name: idx + 1 for idx, name in enumerate(tech_cols) if name}

        mandatory_fields = extract_mandatory_fields(ws, col_to_idx)

        current_row = 9
        rows_with_missing_mandatory = set()

        for rec in processed_records:
            row_data = rec['ecc_row']

            mapped_values = {}
            mapped_values['BUKRS'] = rec['s4_cocd']

            if sheet_name in ['Master Details', 'Allocations', 'Origin', 'Depreciation Areas',
                             'Cumulative Values', 'Posted Values', 'Posting Information',
                             'Time-Dependent Data', 'Acct. Assignmt. for Investment']:
                mapped_values['ANLN1'] = rec['s4_asset']
                mapped_values['ANLN2'] = rec['s4_subnumber']

            if sheet_name == 'Master Details':
                mapped_values['ANLKL'] = clean_string(row_data.get('S4 asset class '))
                mapped_values['TXT50'] = clean_string(row_data.get('Description'))
                mapped_values['TXA50_MORE'] = clean_string(row_data.get('Description'))
                mapped_values['SERNR'] = clean_string(row_data.get('Serial no.'))
                mapped_values['INVNR'] = clean_string(row_data.get('Inventory number'))
                mapped_values['MAIN_DESCRIPT'] = clean_string(row_data.get('Description'))

            elif sheet_name == 'Allocations':
                ins_repl = clean_string(row_data.get('Insurable for replacement Value (Yes or No)'))
                mapped_values['EVALGROUP1'] = ins_repl

            elif sheet_name == 'Origin':
                mapped_values['ORIG_VALUE'] = clean_float(row_data.get('Acquisition Value'))

            elif sheet_name == 'Depreciation Areas':
                mapped_values['AFABE'] = "01"
                mapped_values['AFASL'] = "LINS"
                mapped_values['NDJAR'] = clean_int(row_data.get('Use'), default=None)
                mapped_values['NDPER'] = clean_int(row_data.get('Per'), default=None)
                mapped_values['NDABJ'] = clean_int(row_data.get('EUL'), default=None)
                mapped_values['NDABP'] = clean_int(row_data.get('ELP'), default=None)
                mapped_values['AFABG'] = clean_date(row_data.get('ODep.Start'))

            elif sheet_name == 'Cumulative Values':
                mapped_values['AFABE'] = "01"
                mapped_values['GJAHR'] = clean_int(row_data.get('Year'))
                mapped_values['KANSW'] = clean_float(row_data.get('Acquisition Value'))
                mapped_values['KNAFA'] = clean_float(row_data.get('Prior_Year_Accum_Dep'))
                mapped_values['KAAFA'] = clean_float(row_data.get('Pstd.unpl.dep.'))
                mapped_values['KMAFA'] = clean_float(row_data.get('Transf.Acq.Value'))
                mapped_values['KAUFN'] = clean_float(row_data.get('Transf.Accu.Dep'))

            elif sheet_name == 'Posted Values':
                mapped_values['AFABE'] = "01"
                mapped_values['GJAHR'] = clean_int(row_data.get('Year'))
                mapped_values['NAFAG'] = clean_float(row_data.get('Posted.Dep.CY'))
                mapped_values['AAFAG'] = clean_float(row_data.get('Pstd.unpl.dep.'))
                mapped_values['MAFAG'] = clean_float(row_data.get('Transf.Acq.Value'))
                mapped_values['AUFNG'] = clean_float(row_data.get('Transf.Accu.Dep'))
                mapped_values['LAST_POSTED_DEPR_PERIOD'] = clean_int(row_data.get('Last Dep Per'))

            elif sheet_name == 'Posting Information':
                mapped_values['AKTIV'] = clean_date(row_data.get('Cap.date'))
                mapped_values['DEAKT'] = clean_date(row_data.get('Deact.Date'))

            elif sheet_name == 'Time-Dependent Data':
                mapped_values['GSBER'] = clean_string(row_data.get('BusA'))
                mapped_values['KOSTL'] = rec['s4_cost_center']
                mapped_values['WERKS'] = rec['s4_plant']
                mapped_values['STORT'] = rec['s4_location']

            elif sheet_name == 'Acct. Assignmt. for Investment':
                mapped_values['INVEST_ORD'] = clean_string(row_data.get('Inv. order'))

            row_has_missing = rec['cost_center_missing']

            for field_tech, field_label in mandatory_fields.items():
                if is_blank(mapped_values.get(field_tech)):
                    row_has_missing = True
                    validation_errors.append({
                        'sheet': sheet_name,
                        'field': field_tech,
                        'field_label': field_label,
                        'source_row': rec['source_row'],
                        'source_sheet': rec['source_sheet'],
                        'asset': rec['s4_asset'],
                        'subnumber': rec['s4_subnumber'],
                        'company_code': rec['s4_cocd'],
                    })

            if row_has_missing:
                rows_with_missing_mandatory.add(current_row)

            for col_name, value in mapped_values.items():
                if col_name in col_to_idx:
                    c_idx = col_to_idx[col_name]
                    ws.cell(row=current_row, column=c_idx, value=value)

            current_row += 1

        last_col = len(tech_cols) if tech_cols else ws.max_column
        for flagged_row in rows_with_missing_mandatory:
            for c_idx in range(1, last_col + 1):
                ws.cell(row=flagged_row, column=c_idx).fill = MISSING_MANDATORY_FILL

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf, validation_errors