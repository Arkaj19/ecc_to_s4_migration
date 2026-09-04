# import io
# import datetime
# from copy import copy

# import pandas as pd
# import openpyxl
# from openpyxl.styles import PatternFill


# COMPANY_CODE_MAPPING = {
#     "US01": "1000",
#     "US06": "1001",
#     "CA01": "1200",
# }


# # Expected currency for each S/4 company code
# COMPANY_CODE_CURRENCY_MAPPING = {
#     "1000": "USD",
#     "1001": "USD",
#     "1200": "CAD",
# }


# REASON_CODE_MAPPING = {
#     "DIC": "048",
#     "FRW": "031",
#     "PEC": "021",
#     "PRC": "024",
#     "SSC": "036",
#     "UDC": "001",
#     "UPC": "011",
# }


# def clean_string(value):
#     if value is None or pd.isna(value):
#         return ""

#     if isinstance(value, float) and value.is_integer():
#         return str(int(value))

#     return str(value).strip()


# def clean_float(value, default=None):
#     if value is None or pd.isna(value):
#         return default

#     try:
#         return float(value)
#     except (ValueError, TypeError):
#         return default


# def clean_date(value):
#     if value is None or pd.isna(value):
#         return None

#     if isinstance(value, (datetime.date, datetime.datetime)):
#         return value.date() if isinstance(value, datetime.datetime) else value

#     value_string = str(value).strip()

#     if value_string in ("", "00/00/0000", "00.00.0000", "NaT"):
#         return None

#     for date_format in (
#         "%Y-%m-%d %H:%M:%S",
#         "%Y-%m-%d",
#         "%m/%d/%Y",
#         "%d/%m/%Y",
#         "%d.%m.%Y",
#         "%Y%m%d",
#     ):
#         try:
#             return datetime.datetime.strptime(
#                 value_string,
#                 date_format
#             ).date()
#         except ValueError:
#             continue

#     return value


# def get_s4_company_code(ecc_company_code):
#     return COMPANY_CODE_MAPPING.get(
#         clean_string(ecc_company_code).upper(),
#         "",
#     )


# def get_reason_code(reason_code):
#     return REASON_CODE_MAPPING.get(
#         clean_string(reason_code).upper(),
#         "",
#     )


# def normalize_amount(amount, debit_credit_indicator):
#     """
#     S -> positive
#     H -> negative

#     Positive values are written normally (500), not as '+500'.
#     """
#     amount = clean_float(amount)

#     if amount is None:
#         return None

#     amount = abs(amount)
#     indicator = clean_string(debit_credit_indicator).upper()

#     if indicator == "S":
#         return amount

#     if indicator == "H":
#         return -amount

#     return amount


# def get_reference_value(reference):
#     reference = clean_string(reference)

#     return reference if reference else "No reference in ECC"


# def get_document_type_mappings(
#     document_type,
#     assignment,
#     text,
#     reference,
# ):
#     document_type = clean_string(document_type).upper()
#     assignment = clean_string(assignment)
#     text = clean_string(text)
#     reference = get_reference_value(reference)

#     if document_type == "RV":
#         return {
#             "reference_document_number": assignment,
#             "assignment": "",
#         }

#     if document_type == "DZ":
#         return {
#             "reference_document_number": reference,
#             "assignment": text,
#         }

#     return {
#         "reference_document_number": assignment,
#         "assignment": assignment,
#     }


# class RegistryMismatchError(ValueError):
#     pass


# class CurrencyReviewRequiredError(ValueError):
#     """
#     Raised when company-code/currency mismatches are found and
#     the user has not yet selected KEEP or DELETE.
#     """

#     def __init__(self, review_payload):
#         self.review_payload = review_payload

#         super().__init__(
#             "Company code/currency mismatches require a user decision."
#         )


# REQUIRED_AR_COLUMNS = [
#     "Company Code",
#     "Customer",
#     "Assignment",
#     "Document Number",
#     "Document Date",
#     "Currency",
#     "Reference",
#     "Document Type",
#     "Debit/Credit Ind.",
#     "Amount",
# ]


# def process_ar_registry(
#     registry_file,
#     template_path="templates/Merged File all DOC Types.xlsx",
#     currency_action=None,
# ):
#     """
#     Processes the ECC Accounts Receivable registry and populates
#     the S/4 migration template.

#     Currency review workflow:

#         1. Process the registry normally.
#         2. Check S/4 company code against currency.
#         3. If mismatches exist and currency_action is None:
#            stop and return a review payload.
#         4. KEEP:
#            retain mismatched rows and highlight them red.
#         5. DELETE:
#            remove mismatched rows from the main sheet and place
#            them into a "Currency Mismatch Dump" sheet.

#     Existing AR mapping rules remain unchanged.

#     The target Document Type is always UE.

#     Explicit business rules:
#       - RV: Assignment -> Reference Document Number; Assignment blank
#       - DZ: Text -> Assignment; Reference -> Reference Document Number
#       - Other: Assignment -> Reference Document Number and Assignment
#       - Empty Reference -> "No reference in ECC"
#       - S -> positive Amount
#       - H -> negative Amount
#       - Positive amounts are written without a '+' sign
#       - Reason codes use REASON_CODE_MAPPING
#     """

#     # ------------------------------------------------------------
#     # Validate currency action
#     # ------------------------------------------------------------

#     if currency_action is not None:
#         currency_action = clean_string(currency_action).upper()

#         if currency_action not in {"KEEP", "DELETE"}:
#             raise RegistryMismatchError(
#                 "Invalid currency_action. Expected KEEP or DELETE."
#             )

#     # ------------------------------------------------------------
#     # Read ECC registry
#     # ------------------------------------------------------------

#     df = pd.read_excel(registry_file)

#     if df.empty:
#         raise RegistryMismatchError(
#             "The uploaded AR registry is empty."
#         )

#     # ------------------------------------------------------------
#     # Validate required columns
#     # ------------------------------------------------------------

#     missing_columns = [
#         column
#         for column in REQUIRED_AR_COLUMNS
#         if column not in df.columns
#     ]

#     if missing_columns:
#         raise RegistryMismatchError(
#             "The uploaded file does not contain the required AR "
#             f"column(s): {', '.join(missing_columns)}."
#         )

#     # ------------------------------------------------------------
#     # Load migration template
#     # ------------------------------------------------------------

#     wb = openpyxl.load_workbook(template_path)

#     # Prefer Customer Open Items sheet.
#     # Otherwise use first worksheet.
#     if "Customer Open Items" in wb.sheetnames:
#         ws = wb["Customer Open Items"]
#     else:
#         ws = wb[wb.sheetnames[0]]

#     # ------------------------------------------------------------
#     # Find technical target field columns
#     # ------------------------------------------------------------

#     technical_columns = {}

#     # Technical headers are normally in row 5.
#     for column in range(1, ws.max_column + 1):
#         value = clean_string(
#             ws.cell(row=5, column=column).value
#         )

#         if value:
#             technical_columns[value] = column

#     # Fallback if technical headers are in row 1.
#     if not technical_columns:
#         for column in range(1, ws.max_column + 1):
#             value = clean_string(
#                 ws.cell(row=1, column=column).value
#             )

#             if value:
#                 technical_columns[value] = column

#     # ------------------------------------------------------------
#     # Data starts from row 9
#     # ------------------------------------------------------------

#     data_start_row = 9

#     # ------------------------------------------------------------
#     # Remove existing/example data while preserving formatting
#     # ------------------------------------------------------------

#     for row in range(data_start_row, ws.max_row + 1):
#         for column in range(1, ws.max_column + 1):
#             ws.cell(
#                 row=row,
#                 column=column
#             ).value = None

#     # ------------------------------------------------------------
#     # Currency mismatch collection
#     # ------------------------------------------------------------

#     currency_mismatches = []

#     # Store mapped rows so that DELETE can compact the main sheet
#     # cleanly and KEEP can highlight the correct rows.
#     mapped_rows = []

#     current_row = data_start_row

#     # ------------------------------------------------------------
#     # Process every ECC row
#     # ------------------------------------------------------------

#     for idx, source_row in df.iterrows():

#         ecc_company_code = clean_string(
#             source_row.get("Company Code")
#         )

#         s4_company_code = get_s4_company_code(
#             ecc_company_code
#         )

#         original_document_type = clean_string(
#             source_row.get("Document Type")
#         ).upper()

#         # --------------------------------------------------------
#         # Existing document type mapping
#         # --------------------------------------------------------

#         document_type_mappings = get_document_type_mappings(
#             document_type=original_document_type,
#             assignment=source_row.get("Assignment"),
#             text=source_row.get("Text"),
#             reference=source_row.get("Reference"),
#         )

#         # --------------------------------------------------------
#         # Existing amount mapping
#         # --------------------------------------------------------

#         amount = normalize_amount(
#             source_row.get("Amount"),
#             source_row.get("Debit/Credit Ind."),
#         )

#         # --------------------------------------------------------
#         # Existing tax-code mapping
#         # --------------------------------------------------------

#         if s4_company_code in ("1000", "1001"):
#             tax_code = "I0"

#         elif s4_company_code == "1200":
#             tax_code = "C0"

#         else:
#             tax_code = ""

#         # --------------------------------------------------------
#         # Existing AR mapped values
#         # --------------------------------------------------------

#         mapped_values = {
#             "BUKRS": s4_company_code,

#             "XBLNR": document_type_mappings[
#                 "reference_document_number"
#             ],

#             "KUNNR": clean_string(
#                 source_row.get("Customer")
#             ),

#             "GKONT": "9999900000",

#             # Final target document type is always UE.
#             "BLART": "UE",

#             "BLDAT": clean_date(
#                 source_row.get("Document Date")
#             ),

#             "SGTXT": clean_string(
#                 source_row.get("Text")
#             ),

#             "WAERS": clean_string(
#                 source_row.get("Currency")
#             ),

#             "WRBTR": amount,

#             "MWSKZ": tax_code,

#             "ZTERM": clean_string(
#                 source_row.get("Terms of Payment")
#             ),

#             "ZFBDT": clean_date(
#                 source_row.get("Baseline Payment Dte")
#             ),

#             "ZBD1T": clean_string(
#                 source_row.get("Days 1")
#             ),

#             "ZBD1P": clean_float(
#                 source_row.get("Disc.percent 1")
#             ),

#             "ZBD2T": clean_string(
#                 source_row.get("Days 2")
#             ),

#             "ZBD2P": clean_float(
#                 source_row.get("Disc.percent 2")
#             ),

#             "ZBD3T": clean_string(
#                 source_row.get("Days Net")
#             ),

#             "SKFBT": clean_float(
#                 source_row.get("Discount base")
#             ),

#             "KKBER": clean_string(
#                 source_row.get("Credit Control Area")
#             ),

#             "ZUONR": document_type_mappings[
#                 "assignment"
#             ],

#             "RSTGR": get_reason_code(
#                 source_row.get("Reason code")
#             ),
#         }

#         # --------------------------------------------------------
#         # NEW: Company Code / Currency mismatch check
#         # --------------------------------------------------------

#         expected_currency = COMPANY_CODE_CURRENCY_MAPPING.get(
#             clean_string(s4_company_code).upper()
#         )

#         actual_currency = clean_string(
#             mapped_values.get("WAERS")
#         ).upper()

#         is_currency_mismatch = (
#             expected_currency is not None
#             and actual_currency != expected_currency
#         )

#         if is_currency_mismatch:

#             currency_mismatches.append({
#                 "source_row": int(idx) + 2,
#                 "excel_row": current_row,
#                 "company_code": clean_string(
#                     s4_company_code
#                 ),
#                 "currency": actual_currency,
#                 "expected_currency": expected_currency,
#                 "reason": (
#                     f"Company code "
#                     f"{clean_string(s4_company_code)} "
#                     f"expects {expected_currency}, "
#                     f"but the row contains "
#                     f"{actual_currency or 'blank'}."
#                 ),
#             })

#         # --------------------------------------------------------
#         # Write mapped values into the migration sheet
#         # --------------------------------------------------------

#         for technical_field, value in mapped_values.items():

#             if technical_field in technical_columns:

#                 ws.cell(
#                     row=current_row,
#                     column=technical_columns[technical_field],
#                     value=value,
#                 )

#         mapped_rows.append({
#             "excel_row": current_row,
#             "mapped_values": mapped_values,
#             "is_currency_mismatch": is_currency_mismatch,
#         })

#         current_row += 1

#     # ============================================================
#     # IMPORTANT:
#     # DO NOT RETURN MIGRATION FILE YET IF REVIEW IS REQUIRED
#     # ============================================================

#     if currency_mismatches and currency_action is None:

#         review_payload = {
#             "status": "CURRENCY_REVIEW_REQUIRED",
#             "message": (
#                 "Company code and currency mismatches were found. "
#                 "Choose KEEP or DELETE before migration file generation."
#             ),
#             "mismatch_count": len(currency_mismatches),
#             "options": [
#                 "KEEP",
#                 "DELETE",
#             ],
#             "mismatches": currency_mismatches,
#         }

#         raise CurrencyReviewRequiredError(
#             review_payload
#         )

#     # ============================================================
#     # KEEP
#     # ============================================================

#     if currency_mismatches and currency_action == "KEEP":

#         red_fill = PatternFill(
#             fill_type="solid",
#             fgColor="FFC7CE",
#         )

#         for mismatch in currency_mismatches:

#             excel_row = mismatch["excel_row"]

#             for column in range(
#                 1,
#                 ws.max_column + 1
#             ):
#                 ws.cell(
#                     row=excel_row,
#                     column=column
#                 ).fill = copy(red_fill)

#     # ============================================================
#     # DELETE
#     # ============================================================

#     dump_rows_count = 0

#     if currency_mismatches and currency_action == "DELETE":

#         # --------------------------------------------------------
#         # Create dump sheet from the same template structure.
#         # --------------------------------------------------------

#         dump_sheet_name = "Currency Mismatch Dump"

#         # Remove existing dump sheet if somehow present.
#         if dump_sheet_name in wb.sheetnames:
#             del wb[dump_sheet_name]

#         dump_ws = wb.copy_worksheet(ws)
#         dump_ws.title = dump_sheet_name

#         # --------------------------------------------------------
#         # Clear all data rows in dump sheet first.
#         # --------------------------------------------------------

#         for row in range(
#             data_start_row,
#             dump_ws.max_row + 1
#         ):
#             for column in range(
#                 1,
#                 dump_ws.max_column + 1
#             ):
#                 dump_ws.cell(
#                     row=row,
#                     column=column
#                 ).value = None

#         # --------------------------------------------------------
#         # Identify rows to dump
#         # --------------------------------------------------------

#         mismatch_source_rows = {
#             mismatch["excel_row"]
#             for mismatch in currency_mismatches
#         }

#         # --------------------------------------------------------
#         # Copy mismatched rows to dump sheet
#         # --------------------------------------------------------

#         dump_row = data_start_row

#         for original_row in mapped_rows:

#             if not original_row["is_currency_mismatch"]:
#                 continue

#             source_excel_row = original_row["excel_row"]

#             for column in range(
#                 1,
#                 ws.max_column + 1
#             ):

#                 source_cell = ws.cell(
#                     row=source_excel_row,
#                     column=column
#                 )

#                 target_cell = dump_ws.cell(
#                     row=dump_row,
#                     column=column
#                 )

#                 target_cell.value = source_cell.value

#                 if source_cell.has_style:
#                     target_cell._style = copy(
#                         source_cell._style
#                     )

#                 if source_cell.number_format:
#                     target_cell.number_format = (
#                         source_cell.number_format
#                     )

#                 if source_cell.alignment:
#                     target_cell.alignment = copy(
#                         source_cell.alignment
#                     )

#                 if source_cell.protection:
#                     target_cell.protection = copy(
#                         source_cell.protection
#                     )

#                 if source_cell.font:
#                     target_cell.font = copy(
#                         source_cell.font
#                     )

#                 if source_cell.fill:
#                     target_cell.fill = copy(
#                         source_cell.fill
#                     )

#                 if source_cell.border:
#                     target_cell.border = copy(
#                         source_cell.border
#                     )

#             # Preserve row height if present.
#             if source_excel_row in ws.row_dimensions:
#                 dump_ws.row_dimensions[
#                     dump_row
#                 ].height = ws.row_dimensions[
#                     source_excel_row
#                 ].height

#             dump_row += 1

#         dump_rows_count = (
#             dump_row - data_start_row
#         )

#         # --------------------------------------------------------
#         # Compact retained rows in main sheet.
#         # --------------------------------------------------------

#         retained_rows = [
#             row
#             for row in mapped_rows
#             if not row["is_currency_mismatch"]
#         ]

#         # Clear the entire main data area.
#         for row in range(
#             data_start_row,
#             ws.max_row + 1
#         ):
#             for column in range(
#                 1,
#                 ws.max_column + 1
#             ):
#                 ws.cell(
#                     row=row,
#                     column=column
#                 ).value = None

#         # --------------------------------------------------------
#         # Rewrite only retained rows.
#         # --------------------------------------------------------

#         main_row = data_start_row

#         for retained_row in retained_rows:

#             mapped_values = retained_row[
#                 "mapped_values"
#             ]

#             for technical_field, value in mapped_values.items():

#                 if technical_field in technical_columns:

#                     ws.cell(
#                         row=main_row,
#                         column=technical_columns[
#                             technical_field
#                         ],
#                         value=value,
#                     )

#             main_row += 1

#     # ============================================================
#     # Prepare output
#     # ============================================================

#     output = io.BytesIO()

#     wb.save(output)

#     output.seek(0)

#     # ============================================================
#     # Attach review metadata for main.py
#     # ============================================================

#     retained_row_count = len(mapped_rows)

#     if currency_action == "DELETE":
#         retained_row_count = (
#             len(mapped_rows)
#             - len(currency_mismatches)
#         )

#     output.currency_review = {
#         "status": "COMPLETED",
#         "action": currency_action,
#         "mismatch_count": len(currency_mismatches),
#         "dump_rows": dump_rows_count,
#         "retained_rows": retained_row_count,
#     }

#     # main.py expects:
#     #     out_buf, validation_errors
#     #
#     # This processor does not currently generate validation
#     # errors, so return an empty list.
#     validation_errors = []

#     return output, validation_errors
import io
import datetime
from copy import copy
import time

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


COMPANY_CODE_MAPPING = {
    "US01": "1000",
    "US06": "1001",
    "CA01": "1200",
}


# Expected currency for each S/4 company code
COMPANY_CODE_CURRENCY_MAPPING = {
    "1000": "USD",
    "1001": "USD",
    "1200": "CAD",
}


REASON_CODE_MAPPING = {
    "DIC": "048",
    "FRW": "031",
    "PEC": "021",
    "PRC": "024",
    "SSC": "036",
    "UDC": "001",
    "UPC": "011",
}


def clean_string(value):
    if value is None or pd.isna(value):
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def clean_float(value, default=None):
    if value is None or pd.isna(value):
        return default

    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def clean_date(value):
    if value is None or pd.isna(value):
        return None

    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.date() if isinstance(value, datetime.datetime) else value

    value_string = str(value).strip()

    if value_string in ("", "00/00/0000", "00.00.0000", "NaT"):
        return None

    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%d.%m.%Y",
        "%Y%m%d",
    ):
        try:
            return datetime.datetime.strptime(
                value_string,
                date_format
            ).date()
        except ValueError:
            continue

    return value


def get_s4_company_code(ecc_company_code):
    return COMPANY_CODE_MAPPING.get(
        clean_string(ecc_company_code).upper(),
        "",
    )


def get_reason_code(reason_code):
    return REASON_CODE_MAPPING.get(
        clean_string(reason_code).upper(),
        "",
    )


def normalize_amount(amount, debit_credit_indicator):
    """
    S -> positive
    H -> negative

    Positive values are written normally (500), not as '+500'.
    """
    amount = clean_float(amount)

    if amount is None:
        return None

    amount = abs(amount)
    indicator = clean_string(debit_credit_indicator).upper()

    if indicator == "S":
        return amount

    if indicator == "H":
        return -amount

    return amount


def get_reference_value(reference):
    reference = clean_string(reference)

    return reference if reference else "No reference in ECC"


def get_document_type_mappings(
    document_type,
    assignment,
    text,
    reference,
):
    document_type = clean_string(document_type).upper()
    assignment = clean_string(assignment)
    text = clean_string(text)
    reference = get_reference_value(reference)

    if document_type == "RV":
        return {
            "reference_document_number": assignment,
            "assignment": "",
        }

    if document_type == "DZ":
        return {
            "reference_document_number": reference,
            "assignment": text,
        }

    return {
        "reference_document_number": assignment,
        "assignment": assignment,
    }


class RegistryMismatchError(ValueError):
    pass


class CurrencyReviewRequiredError(ValueError):
    """
    Raised when company-code/currency mismatches are found and
    the user has not yet selected KEEP or DELETE.
    """

    def __init__(self, review_payload):
        self.review_payload = review_payload

        super().__init__(
            "Company code/currency mismatches require a user decision."
        )


REQUIRED_AR_COLUMNS = [
    "Company Code",
    "Customer",
    "Assignment",
    "Document Number",
    "Document Date",
    "Currency",
    "Reference",
    "Document Type",
    "Debit/Credit Ind.",
    "Amount",
]


def process_ar_registry(
    registry_file,
    template_path="templates/Merged File all DOC Types.xlsx",
    currency_action=None,
):
    """
    Processes the ECC Accounts Receivable registry and populates
    the S/4 migration template.

    Currency review workflow:

        1. Process the registry normally.
        2. Check S/4 company code against currency.
        3. If mismatches exist and currency_action is None:
           stop and return a review payload.
        4. KEEP:
           retain mismatched rows and highlight them red.
        5. DELETE:
           remove mismatched rows from the main sheet and place
           them into a "Currency Mismatch Dump" sheet.

    Existing AR mapping rules remain unchanged.

    The target Document Type is always UE.

    Explicit business rules:
      - RV: Assignment -> Reference Document Number; Assignment blank
      - DZ: Text -> Assignment; Reference -> Reference Document Number
      - Other: Assignment -> Reference Document Number and Assignment
      - Empty Reference -> "No reference in ECC"
      - S -> positive Amount
      - H -> negative Amount
      - Positive amounts are written without a '+' sign
      - Reason codes use REASON_CODE_MAPPING
    """

    start_time = time.time()
    print(f"AR Processing started at {start_time:.2f}s")

    # ------------------------------------------------------------
    # Validate currency action
    # ------------------------------------------------------------

    if currency_action is not None:
        currency_action = clean_string(currency_action).upper()

        if currency_action not in {"KEEP", "DELETE"}:
            raise RegistryMismatchError(
                "Invalid currency_action. Expected KEEP or DELETE."
            )

    # ------------------------------------------------------------
    # Read ECC registry
    # ------------------------------------------------------------

    df = pd.read_excel(registry_file)

    if df.empty:
        raise RegistryMismatchError(
            "The uploaded AR registry is empty."
        )

    # ------------------------------------------------------------
    # Validate required columns
    # ------------------------------------------------------------

    missing_columns = [
        column
        for column in REQUIRED_AR_COLUMNS
        if column not in df.columns
    ]

    if missing_columns:
        raise RegistryMismatchError(
            "The uploaded file does not contain the required AR "
            f"column(s): {', '.join(missing_columns)}."
        )

    # ------------------------------------------------------------
    # Load migration template with optimized settings
    # ------------------------------------------------------------

    print(f"Loading template at {time.time() - start_time:.2f}s")
    wb = openpyxl.load_workbook(template_path)
    
    # Disable auto-calculation for performance
    try:
        wb.calculation.calcMode = 'manual'
    except AttributeError:
        pass  # Some versions don't support this

    # Prefer Customer Open Items sheet.
    # Otherwise use first worksheet.
    if "Customer Open Items" in wb.sheetnames:
        ws = wb["Customer Open Items"]
    else:
        ws = wb[wb.sheetnames[0]]

    # ------------------------------------------------------------
    # Find technical target field columns
    # ------------------------------------------------------------

    technical_columns = {}

    # Technical headers are normally in row 5.
    for column in range(1, ws.max_column + 1):
        value = clean_string(
            ws.cell(row=5, column=column).value
        )

        if value:
            technical_columns[value] = column

    # Fallback if technical headers are in row 1.
    if not technical_columns:
        for column in range(1, ws.max_column + 1):
            value = clean_string(
                ws.cell(row=1, column=column).value
            )

            if value:
                technical_columns[value] = column

    # ------------------------------------------------------------
    # Data starts from row 9
    # ------------------------------------------------------------

    data_start_row = 9

    # ------------------------------------------------------------
    # Remove existing/example data while preserving formatting
    # ------------------------------------------------------------

    print(f"Clearing existing data at {time.time() - start_time:.2f}s")
    for row in range(data_start_row, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            ws.cell(
                row=row,
                column=column
            ).value = None

    # ------------------------------------------------------------
    # Currency mismatch collection
    # ------------------------------------------------------------

    currency_mismatches = []

    # Store mapped rows so that DELETE can compact the main sheet
    # cleanly and KEEP can highlight the correct rows.
    mapped_rows = []

    current_row = data_start_row

    # ------------------------------------------------------------
    # Process every ECC row
    # ------------------------------------------------------------

    print(f"Processing {len(df)} rows at {time.time() - start_time:.2f}s")
    
    for idx, source_row in df.iterrows():

        ecc_company_code = clean_string(
            source_row.get("Company Code")
        )

        s4_company_code = get_s4_company_code(
            ecc_company_code
        )

        original_document_type = clean_string(
            source_row.get("Document Type")
        ).upper()

        # --------------------------------------------------------
        # Existing document type mapping
        # --------------------------------------------------------

        document_type_mappings = get_document_type_mappings(
            document_type=original_document_type,
            assignment=source_row.get("Assignment"),
            text=source_row.get("Text"),
            reference=source_row.get("Reference"),
        )

        # --------------------------------------------------------
        # Existing amount mapping
        # --------------------------------------------------------

        amount = normalize_amount(
            source_row.get("Amount"),
            source_row.get("Debit/Credit Ind."),
        )

        # --------------------------------------------------------
        # Existing tax-code mapping
        # --------------------------------------------------------

        if s4_company_code in ("1000", "1001"):
            tax_code = "I0"

        elif s4_company_code == "1200":
            tax_code = "C0"

        else:
            tax_code = ""

        # --------------------------------------------------------
        # Existing AR mapped values
        # --------------------------------------------------------

        mapped_values = {
            "BUKRS": s4_company_code,

            "XBLNR": document_type_mappings[
                "reference_document_number"
            ],

            "KUNNR": clean_string(
                source_row.get("Customer")
            ),

            "GKONT": "9999900000",

            # Final target document type is always UE.
            "BLART": "UE",

            "BLDAT": clean_date(
                source_row.get("Document Date")
            ),

            "SGTXT": clean_string(
                source_row.get("Text")
            ),

            "WAERS": clean_string(
                source_row.get("Currency")
            ),

            "WRBTR": amount,

            "MWSKZ": tax_code,

            "ZTERM": clean_string(
                source_row.get("Terms of Payment")
            ),

            "ZFBDT": clean_date(
                source_row.get("Baseline Payment Dte")
            ),

            "ZBD1T": clean_string(
                source_row.get("Days 1")
            ),

            "ZBD1P": clean_float(
                source_row.get("Disc.percent 1")
            ),

            "ZBD2T": clean_string(
                source_row.get("Days 2")
            ),

            "ZBD2P": clean_float(
                source_row.get("Disc.percent 2")
            ),

            "ZBD3T": clean_string(
                source_row.get("Days Net")
            ),

            "SKFBT": clean_float(
                source_row.get("Discount base")
            ),

            "KKBER": clean_string(
                source_row.get("Credit Control Area")
            ),

            "ZUONR": document_type_mappings[
                "assignment"
            ],

            "RSTGR": get_reason_code(
                source_row.get("Reason code")
            ),
        }

        # --------------------------------------------------------
        # NEW: Company Code / Currency mismatch check
        # --------------------------------------------------------

        expected_currency = COMPANY_CODE_CURRENCY_MAPPING.get(
            clean_string(s4_company_code).upper()
        )

        actual_currency = clean_string(
            mapped_values.get("WAERS")
        ).upper()

        is_currency_mismatch = (
            expected_currency is not None
            and actual_currency != expected_currency
        )

        if is_currency_mismatch:

            currency_mismatches.append({
                "source_row": int(idx) + 2,
                "excel_row": current_row,
                "company_code": clean_string(
                    s4_company_code
                ),
                "currency": actual_currency,
                "expected_currency": expected_currency,
                "reason": (
                    f"Company code "
                    f"{clean_string(s4_company_code)} "
                    f"expects {expected_currency}, "
                    f"but the row contains "
                    f"{actual_currency or 'blank'}."
                ),
            })

        # --------------------------------------------------------
        # Write mapped values into the migration sheet
        # --------------------------------------------------------

        for technical_field, value in mapped_values.items():

            if technical_field in technical_columns:

                ws.cell(
                    row=current_row,
                    column=technical_columns[technical_field],
                    value=value,
                )

        mapped_rows.append({
            "excel_row": current_row,
            "mapped_values": mapped_values,
            "is_currency_mismatch": is_currency_mismatch,
        })

        current_row += 1

    print(f"Rows processed at {time.time() - start_time:.2f}s")
    print(f"Currency mismatches found: {len(currency_mismatches)}")

    # ============================================================
    # IMPORTANT:
    # DO NOT RETURN MIGRATION FILE YET IF REVIEW IS REQUIRED
    # ============================================================

    if currency_mismatches and currency_action is None:

        review_payload = {
            "status": "CURRENCY_REVIEW_REQUIRED",
            "message": (
                "Company code and currency mismatches were found. "
                "Choose KEEP or DELETE before migration file generation."
            ),
            "mismatch_count": len(currency_mismatches),
            "options": [
                "KEEP",
                "DELETE",
            ],
            "mismatches": currency_mismatches,
        }

        raise CurrencyReviewRequiredError(
            review_payload
        )

    # ============================================================
    # KEEP - OPTIMIZED VERSION
    # ============================================================

    if currency_mismatches and currency_action == "KEEP":

        keep_start = time.time()
        print(f"KEEP: Starting at {time.time() - start_time:.2f}s")
        print(f"KEEP: Highlighting {len(currency_mismatches)} rows")
        
        red_fill = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE",
        )

        # ✅ OPTIMIZATION: Collect all rows that need highlighting
        rows_to_highlight = [
            mismatch["excel_row"] 
            for mismatch in currency_mismatches
        ]

        # ✅ OPTIMIZATION: Use set for faster lookup
        rows_to_highlight_set = set(rows_to_highlight)

        # ✅ OPTIMIZATION: Highlight all mismatched rows in one pass
        # Instead of iterating through mismatches and then through columns,
        # we iterate through rows that need highlighting and apply fill
        for excel_row in rows_to_highlight_set:
            for column in range(1, ws.max_column + 1):
                ws.cell(
                    row=excel_row,
                    column=column
                ).fill = copy(red_fill)

        print(f"KEEP: Completed at {time.time() - start_time:.2f}s (took {time.time() - keep_start:.2f}s)")

    # ============================================================
    # DELETE - OPTIMIZED VERSION
    # ============================================================

    dump_rows_count = 0

    if currency_mismatches and currency_action == "DELETE":

        delete_start = time.time()
        print(f"DELETE: Starting at {time.time() - start_time:.2f}s")

        # --------------------------------------------------------
        # OPTIMIZATION 1: Create dump sheet WITHOUT copying all data
        # --------------------------------------------------------

        dump_sheet_name = "Currency Mismatch Dump"

        # Remove existing dump sheet if somehow present.
        if dump_sheet_name in wb.sheetnames:
            del wb[dump_sheet_name]

        # ✅ Create a new sheet instead of copying the entire worksheet
        dump_ws = wb.create_sheet(dump_sheet_name)

        # ✅ Copy ONLY the header rows (rows 1-8) and structure
        header_rows_to_copy = min(8, ws.max_row)
        
        for row_idx in range(1, header_rows_to_copy + 1):
            for col_idx in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=row_idx, column=col_idx)
                target_cell = dump_ws.cell(row=row_idx, column=col_idx)
                target_cell.value = source_cell.value
                
                # ✅ Only copy styles for header cells (fewer cells)
                if source_cell.has_style:
                    # Only copy if style is not default
                    if source_cell.font and source_cell.font.name:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.fill and source_cell.fill.patternType:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format

        # ✅ Copy column widths from source sheet
        for col_idx in range(1, ws.max_column + 1):
            if col_idx in ws.column_dimensions:
                dump_ws.column_dimensions[col_idx].width = ws.column_dimensions[col_idx].width

        # --------------------------------------------------------
        # OPTIMIZATION 2: Collect mismatched rows efficiently
        # --------------------------------------------------------

        mismatched_rows_data = [
            row for row in mapped_rows 
            if row["is_currency_mismatch"]
        ]

        print(f"DELETE: Found {len(mismatched_rows_data)} mismatched rows at {time.time() - start_time:.2f}s")

        # --------------------------------------------------------
        # OPTIMIZATION 3: Batch write mismatched rows to dump sheet
        # --------------------------------------------------------

        dump_row = data_start_row

        for row_data in mismatched_rows_data:
            source_excel_row = row_data["excel_row"]
            
            # ✅ Optimize: Only copy values and essential formatting
            for col_idx in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=source_excel_row, column=col_idx)
                target_cell = dump_ws.cell(row=dump_row, column=col_idx)
                
                # Copy value
                target_cell.value = source_cell.value
                
                # ✅ Only copy number_format (most important for data integrity)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
            
            # Copy row height if present
            if source_excel_row in ws.row_dimensions:
                dump_ws.row_dimensions[dump_row].height = ws.row_dimensions[source_excel_row].height
            
            dump_row += 1

        dump_rows_count = (dump_row - data_start_row)
        print(f"DELETE: Dump sheet created with {dump_rows_count} rows at {time.time() - start_time:.2f}s")

        # --------------------------------------------------------
        # OPTIMIZATION 4: Compact retained rows efficiently
        # --------------------------------------------------------

        retained_rows = [
            row for row in mapped_rows
            if not row["is_currency_mismatch"]
        ]

        print(f"DELETE: Retaining {len(retained_rows)} rows at {time.time() - start_time:.2f}s")

        # ✅ Clear ONLY the data area that was written
        rows_to_clear = len(mapped_rows)
        
        for row in range(data_start_row, data_start_row + rows_to_clear):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col_idx).value = None

        # ✅ Rewrite retained rows efficiently
        main_row = data_start_row

        for retained_row in retained_rows:
            mapped_values = retained_row["mapped_values"]

            for technical_field, value in mapped_values.items():
                if technical_field in technical_columns:
                    ws.cell(
                        row=main_row,
                        column=technical_columns[technical_field],
                        value=value,
                    )

            main_row += 1

        print(f"DELETE: Completed at {time.time() - start_time:.2f}s (took {time.time() - delete_start:.2f}s)")

    # ============================================================
    # Prepare output
    # ============================================================

    print(f"Saving workbook at {time.time() - start_time:.2f}s")
    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    # ============================================================
    # Attach review metadata for main.py
    # ============================================================

    retained_row_count = len(mapped_rows)

    if currency_action == "DELETE":
        retained_row_count = (
            len(mapped_rows)
            - len(currency_mismatches)
        )

    output.currency_review = {
        "status": "COMPLETED",
        "action": currency_action,
        "mismatch_count": len(currency_mismatches),
        "dump_rows": dump_rows_count,
        "retained_rows": retained_row_count,
    }

    # main.py expects:
    #     out_buf, validation_errors
    #
    # This processor does not currently generate validation
    # errors, so return an empty list.
    validation_errors = []

    print(f"AR Processing completed at {time.time() - start_time:.2f}s total")
    
    return output, validation_errors