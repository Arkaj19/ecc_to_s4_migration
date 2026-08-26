# # import io
# # import pandas as pd
# # import openpyxl
# # import datetime
# # from validation_utils import extract_mandatory_fields, is_blank


# # def clean_string(val):
# #     """
# #     Convert a value to a clean string.

# #     Prevents values such as 1.0 from being written when the
# #     source Excel contains an integer-like float.
# #     """
# #     if pd.isna(val) or val is None:
# #         return ""

# #     if isinstance(val, float) and val.is_integer():
# #         return str(int(val))

# #     return str(val).strip()


# # def clean_float(val, default=None):
# #     """
# #     Convert a value to float.

# #     Blank / NaN values return the supplied default.
# #     """
# #     if pd.isna(val) or val is None:
# #         return default

# #     try:
# #         return float(val)
# #     except (ValueError, TypeError):
# #         return val


# # class RegistryMismatchError(ValueError):
# #     """Raised when the uploaded file doesn't look like a Credit registry."""
# #     pass


# # # CustomerNumber is required because the loop below skips (via
# # # `if not customer_number: continue`) any row where it's missing — so a
# # # file without this column would previously produce a "successful" run
# # # with zero rows written, instead of an error.
# # REQUIRED_CREDIT_COLUMNS = ['CustomerNumber', 'Risk Cat', 'Credit rep.group', 'Credit Limit']


# # def validate_credit_registry(df):
# #     missing = [c for c in REQUIRED_CREDIT_COLUMNS if c not in df.columns]
# #     if missing:
# #         raise RegistryMismatchError(
# #             "This file doesn't look like a Credit registry. "
# #             f"Missing expected column(s): {', '.join(missing)}. "
# #             "Make sure you selected the right file for the Credit Management process."
# #         )


# # def process_credit_registry(
# #     registry_file,
# #     template_path="templates/credit_load_template.xlsx"
# # ) -> io.BytesIO:
# #     """
# #     Processes the Credit Registry spreadsheet and populates:

# #         1. Profile - Credit MD for Cust.
# #         2. Segment - Credit MD for Cust.

# #     Registry -> Credit Profile mapping:

# #         KUNNR              -> CustomerNumber
# #         RUN_ID             -> Sequential
# #         OWN_RATING         -> Blank
# #         CHECK_RULE         -> "01"
# #         LIMIT_RULE         -> "SAP_ALL"
# #         RATING_VAL_DATE    -> Blank
# #         RISK_CLASS         -> Risk Cat
# #         CREDIT_GROUP       -> Credit rep.group

# #     Registry -> Credit Segment mapping:

# #         KUNNR              -> CustomerNumber
# #         RUN_ID             -> Sequential
# #         CREDIT_SEGMENT     -> "1000"
# #         CREDIT_LIMIT       -> Credit Limit

# #     Data is written starting from row 9.
# #     """

# #     # ---------------------------------------------------------
# #     # 1. Read Registry
# #     # ---------------------------------------------------------

# #     df_raw = pd.read_excel(registry_file)

# #     # Fail loudly if this isn't actually a Credit registry, instead of
# #     # silently skipping every row and returning an empty "success".
# #     validate_credit_registry(df_raw)

# #     # ---------------------------------------------------------
# #     # 2. Load Credit Data Load template
# #     # ---------------------------------------------------------

# #     wb = openpyxl.load_workbook(template_path)

# #     sheets_to_fill = [
# #         "Profile - Credit MD for Cust.",
# #         "Segment - Credit MD for Cust."
# #     ]

# #     # ---------------------------------------------------------
# #     # 3. Prepare mapped records
# #     # ---------------------------------------------------------

# #     processed_records = []

# #     run_id = 1

# #     for idx, row in df_raw.iterrows():

# #         customer_number = clean_string(
# #             row.get("CustomerNumber")
# #         )

# #         # Skip rows where CustomerNumber is blank
# #         if not customer_number:
# #             continue

# #         record = {
# #             "customer_number": customer_number,
# #             "run_id": str(run_id),

# #             # +2: pandas index is 0-based and Row 1 is the header, so the
# #             # first data row (index 0) is Excel Row 2.
# #             "source_row": idx + 2,

# #             # Profile fields
# #             "risk_class": clean_string(
# #                 row.get("Risk Cat")
# #             ),

# #             "credit_group": clean_string(
# #                 row.get("Credit rep.group")
# #             ),

# #             # Segment fields
# #             "credit_limit": clean_float(
# #                 row.get("Credit Limit")
# #             )
# #         }

# #         processed_records.append(record)

# #         run_id += 1

# #     # ---------------------------------------------------------
# #     # 4. Populate both sheets
# #     # ---------------------------------------------------------

# #     validation_errors = []

# #     for sheet_name in sheets_to_fill:

# #         if sheet_name not in wb.sheetnames:
# #             print(
# #                 f"Sheet {sheet_name} not found in template, skipping."
# #             )
# #             continue

# #         ws = wb[sheet_name]

# #         # -----------------------------------------------------
# #         # Technical field names are on Row 5
# #         # -----------------------------------------------------

# #         tech_cols = [
# #             clean_string(
# #                 ws.cell(row=5, column=col).value
# #             )
# #             for col in range(1, ws.max_column + 1)
# #         ]

# #         # Build:
# #         # technical_field_name -> Excel column number
# #         col_to_idx = {
# #             name: idx + 1
# #             for idx, name in enumerate(tech_cols)
# #             if name
# #         }

# #         # Which technical fields this sheet marks mandatory (Row 8, '*').
# #         mandatory_fields = extract_mandatory_fields(ws, col_to_idx)

# #         # -----------------------------------------------------
# #         # Data starts from Row 9
# #         # -----------------------------------------------------

# #         current_row = 9

# #         for rec in processed_records:

# #             mapped_values = {}

# #             # =================================================
# #             # PROFILE SHEET
# #             # =================================================

# #             if sheet_name == "Profile - Credit MD for Cust.":

# #                 mapped_values["KUNNR"] = rec["customer_number"]

# #                 mapped_values["RUN_ID"] = rec["run_id"]

# #                 # No mapping provided for OWN_RATING
# #                 mapped_values["OWN_RATING"] = ""

# #                 # Hardcoded for now
# #                 mapped_values["CHECK_RULE"] = "01"

# #                 # Hardcoded
# #                 mapped_values["LIMIT_RULE"] = "SAP_ALL"

# #                 # Blank as specified
# #                 mapped_values["RATING_VAL_DATE"] = ""

# #                 # Risk Cat from Registry
# #                 mapped_values["RISK_CLASS"] = rec["risk_class"]

# #                 # Credit rep.group must remain STRING
# #                 mapped_values["CREDIT_GROUP"] = rec["credit_group"]

# #             # =================================================
# #             # SEGMENT SHEET
# #             # =================================================

# #             elif sheet_name == "Segment - Credit MD for Cust.":

# #                 mapped_values["KUNNR"] = rec["customer_number"]

# #                 mapped_values["RUN_ID"] = rec["run_id"]

# #                 # Hardcoded for now
# #                 mapped_values["CREDIT_SEGMENT"] = "1000"

# #                 mapped_values["CREDIT_LIMIT"] = rec["credit_limit"]

# #                 # Other fields are intentionally left blank
# #                 # because no mapping was provided.

# #             # -------------------------------------------------
# #             # Flag any mandatory field that's blank in the final mapped
# #             # value. The sheet still gets written either way; this only
# #             # records the issue.
# #             # -------------------------------------------------

# #             for field_tech, field_label in mandatory_fields.items():
# #                 if is_blank(mapped_values.get(field_tech)):
# #                     validation_errors.append({
# #                         'sheet': sheet_name,
# #                         'field': field_tech,
# #                         'field_label': field_label,
# #                         'source_row': rec['source_row'],
# #                         'customer_number': rec['customer_number'],
# #                     })

# #             # -------------------------------------------------
# #             # Write mapped values
# #             # -------------------------------------------------

# #             for col_name, value in mapped_values.items():

# #                 if col_name in col_to_idx:

# #                     c_idx = col_to_idx[col_name]

# #                     ws.cell(
# #                         row=current_row,
# #                         column=c_idx,
# #                         value=value
# #                     )

# #             current_row += 1

# #     # ---------------------------------------------------------
# #     # 5. Save workbook to BytesIO
# #     # ---------------------------------------------------------

# #     out_buf = io.BytesIO()

# #     wb.save(out_buf)

# #     out_buf.seek(0)

# #     return out_buf, validation_errors

# import io
# import pandas as pd
# import openpyxl
# import datetime
# from validation_utils import extract_mandatory_fields, is_blank
# from reference_mappings import (
#     load_but_mapping,
#     map_business_partner,
#     load_credit_rep_group_mapping,
#     get_credit_rep_group,
# )


# def clean_string(val):
#     """
#     Convert a value to a clean string.

#     Prevents values such as 1.0 from being written when the
#     source Excel contains an integer-like float.
#     """
#     if pd.isna(val) or val is None:
#         return ""

#     if isinstance(val, float) and val.is_integer():
#         return str(int(val))

#     return str(val).strip()


# def clean_float(val, default=None):
#     """
#     Convert a value to float.

#     Blank / NaN values return the supplied default.
#     """
#     if pd.isna(val) or val is None:
#         return default

#     try:
#         return float(val)
#     except (ValueError, TypeError):
#         return val


# class RegistryMismatchError(ValueError):
#     """Raised when the uploaded file doesn't look like a Credit registry."""
#     pass


# # CustomerNumber is required because the loop below skips (via
# # `if not customer_number: continue`) any row where it's missing — so a
# # file without this column would previously produce a "successful" run
# # with zero rows written, instead of an error.
# REQUIRED_CREDIT_COLUMNS = ['CustomerNumber', 'Risk Cat', 'Credit rep.group', 'Credit Limit']


# def validate_credit_registry(df):
#     missing = [c for c in REQUIRED_CREDIT_COLUMNS if c not in df.columns]
#     if missing:
#         raise RegistryMismatchError(
#             "This file doesn't look like a Credit registry. "
#             f"Missing expected column(s): {', '.join(missing)}. "
#             "Make sure you selected the right file for the Credit Management process."
#         )


# def process_credit_registry(
#     registry_file,
#     template_path="templates/credit_load_template.xlsx",
#     but_path="reference_data/but0id_qs4_500.xlsx",
#     clerk_codes_path="reference_data/DAP_CODES.xlsx"
# ) -> io.BytesIO:
#     """
#     Processes the Credit Registry spreadsheet and populates:

#         1. Profile - Credit MD for Cust.
#         2. Segment - Credit MD for Cust.

#     Registry -> Credit Profile mapping:

#         KUNNR              -> CustomerNumber -> BUT.Business Partner
#         RUN_ID             -> Sequential
#         OWN_RATING         -> Blank
#         CHECK_RULE         -> "01"
#         LIMIT_RULE         -> "SAP_ALL"
#         RATING_VAL_DATE    -> Blank
#         RISK_CLASS         -> Risk Cat
#         CREDIT_GROUP       -> CustomerNumber -> DAP Clerk Codes.Credit Rep Group
#                               (NOT the registry's own Credit rep.group column)

#     Registry -> Credit Segment mapping:

#         KUNNR              -> CustomerNumber -> BUT.Business Partner
#         RUN_ID             -> Sequential
#         CREDIT_SEGMENT     -> "1000"
#         CREDIT_LIMIT       -> Credit Limit

#     Data is written starting from row 9.
#     """

#     # ---------------------------------------------------------
#     # 1. Read Registry
#     # ---------------------------------------------------------

#     df_raw = pd.read_excel(registry_file)

#     # Fail loudly if this isn't actually a Credit registry, instead of
#     # silently skipping every row and returning an empty "success".
#     validate_credit_registry(df_raw)

#     # Customer Number -> Business Partner, via the BUT reference sheet,
#     # scoped to the 'DAP' (customer) Identification Type — see
#     # reference_mappings.load_but_mapping() for why the id_type scoping
#     # matters (the same Identification Number can resolve to a different
#     # Business Partner under 'DAPVEN').
#     customer_but_mapping = load_but_mapping(
#         but_path,
#         id_type="DAP"
#     )

#     # Customer Number -> Credit Rep Group, via the DAP Clerk Codes
#     # reference file. This replaces the registry's own
#     # 'Credit rep.group' column for CREDIT_GROUP.
#     credit_rep_group_mapping = load_credit_rep_group_mapping(
#         clerk_codes_path
#     )

#     # ---------------------------------------------------------
#     # 2. Load Credit Data Load template
#     # ---------------------------------------------------------

#     wb = openpyxl.load_workbook(template_path)

#     sheets_to_fill = [
#         "Profile - Credit MD for Cust.",
#         "Segment - Credit MD for Cust."
#     ]

#     # ---------------------------------------------------------
#     # 3. Prepare mapped records
#     # ---------------------------------------------------------

#     processed_records = []

#     run_id = 1

#     for idx, row in df_raw.iterrows():

#         raw_customer_number = clean_string(
#             row.get("CustomerNumber")
#         )

#         # Skip rows where CustomerNumber is blank
#         if not raw_customer_number:
#             continue

#         record = {
#             # KUNNR in the output is the S/4 Business Partner, resolved
#             # from the registry's raw Customer Number via the BUT sheet
#             # — not the raw registry value itself.
#             "customer_number": map_business_partner(
#                 customer_but_mapping,
#                 raw_customer_number
#             ),

#             # Kept for traceability (validation messages, source lookups)
#             # — this is the key the BUT and Clerk Codes lookups both use.
#             "source_customer_number": raw_customer_number,

#             "run_id": str(run_id),

#             # +2: pandas index is 0-based and Row 1 is the header, so the
#             # first data row (index 0) is Excel Row 2.
#             "source_row": idx + 2,

#             # Profile fields
#             "risk_class": clean_string(
#                 row.get("Risk Cat")
#             ),

#             # Credit Rep Group now comes from the DAP Clerk Codes
#             # reference file, keyed by the raw Customer Number — the
#             # registry's own 'Credit rep.group' column is no longer used
#             # for this field.
#             "credit_group": get_credit_rep_group(
#                 credit_rep_group_mapping,
#                 raw_customer_number
#             ),

#             # Segment fields
#             "credit_limit": clean_float(
#                 row.get("Credit Limit")
#             )
#         }

#         processed_records.append(record)

#         run_id += 1

#     # ---------------------------------------------------------
#     # 4. Populate both sheets
#     # ---------------------------------------------------------

#     validation_errors = []

#     for sheet_name in sheets_to_fill:

#         if sheet_name not in wb.sheetnames:
#             print(
#                 f"Sheet {sheet_name} not found in template, skipping."
#             )
#             continue

#         ws = wb[sheet_name]

#         # -----------------------------------------------------
#         # Technical field names are on Row 5
#         # -----------------------------------------------------

#         tech_cols = [
#             clean_string(
#                 ws.cell(row=5, column=col).value
#             )
#             for col in range(1, ws.max_column + 1)
#         ]

#         # Build:
#         # technical_field_name -> Excel column number
#         col_to_idx = {
#             name: idx + 1
#             for idx, name in enumerate(tech_cols)
#             if name
#         }

#         # Which technical fields this sheet marks mandatory (Row 8, '*').
#         mandatory_fields = extract_mandatory_fields(ws, col_to_idx)

#         # -----------------------------------------------------
#         # Data starts from Row 9
#         # -----------------------------------------------------

#         current_row = 9

#         for rec in processed_records:

#             mapped_values = {}

#             # =================================================
#             # PROFILE SHEET
#             # =================================================

#             if sheet_name == "Profile - Credit MD for Cust.":

#                 mapped_values["KUNNR"] = rec["customer_number"]

#                 mapped_values["RUN_ID"] = rec["run_id"]

#                 # No mapping provided for OWN_RATING
#                 mapped_values["OWN_RATING"] = ""

#                 # Hardcoded for now
#                 mapped_values["CHECK_RULE"] = "01"

#                 # Hardcoded
#                 mapped_values["LIMIT_RULE"] = "SAP_ALL"

#                 # Blank as specified
#                 mapped_values["RATING_VAL_DATE"] = ""

#                 # Risk Cat from Registry
#                 mapped_values["RISK_CLASS"] = rec["risk_class"]

#                 # Credit rep.group must remain STRING
#                 mapped_values["CREDIT_GROUP"] = rec["credit_group"]

#             # =================================================
#             # SEGMENT SHEET
#             # =================================================

#             elif sheet_name == "Segment - Credit MD for Cust.":

#                 mapped_values["KUNNR"] = rec["customer_number"]

#                 mapped_values["RUN_ID"] = rec["run_id"]

#                 # Hardcoded for now
#                 mapped_values["CREDIT_SEGMENT"] = "1000"

#                 mapped_values["CREDIT_LIMIT"] = rec["credit_limit"]

#                 # Other fields are intentionally left blank
#                 # because no mapping was provided.

#             # -------------------------------------------------
#             # Flag any mandatory field that's blank in the final mapped
#             # value. The sheet still gets written either way; this only
#             # records the issue.
#             # -------------------------------------------------

#             for field_tech, field_label in mandatory_fields.items():
#                 if is_blank(mapped_values.get(field_tech)):
#                     validation_errors.append({
#                         'sheet': sheet_name,
#                         'field': field_tech,
#                         'field_label': field_label,
#                         'source_row': rec['source_row'],
#                         'customer_number': rec['customer_number'],
#                         'source_customer_number': rec['source_customer_number'],
#                     })

#             # -------------------------------------------------
#             # Write mapped values
#             # -------------------------------------------------

#             for col_name, value in mapped_values.items():

#                 if col_name in col_to_idx:

#                     c_idx = col_to_idx[col_name]

#                     ws.cell(
#                         row=current_row,
#                         column=c_idx,
#                         value=value
#                     )

#             current_row += 1

#     # ---------------------------------------------------------
#     # 5. Save workbook to BytesIO
#     # ---------------------------------------------------------

#     out_buf = io.BytesIO()

#     wb.save(out_buf)

#     out_buf.seek(0)

#     return out_buf, validation_errors

import io
import pandas as pd
import openpyxl
import datetime
from validation_utils import extract_mandatory_fields, is_blank
from reference_mappings import load_but_mapping, map_business_partner


def clean_string(val):
    """
    Convert a value to a clean string.

    Prevents values such as 1.0 from being written when the
    source Excel contains an integer-like float.
    """
    if pd.isna(val) or val is None:
        return ""

    if isinstance(val, float) and val.is_integer():
        return str(int(val))

    return str(val).strip()


def clean_float(val, default=None):
    """
    Convert a value to float.

    Blank / NaN values return the supplied default.
    """
    if pd.isna(val) or val is None:
        return default

    try:
        return float(val)
    except (ValueError, TypeError):
        return val


def normalize_credit_rep_group(raw_value):
    """
    The Credit Rep Group is a 2-digit code with a leading zero
    ('01', '02', ...). If the registry stores it as a plain number rather
    than text, pandas reads it back as an int/float and the leading zero
    is lost — this restores it so the output matches the expected format.
    Non-numeric values are passed through as a plain string unchanged.
    """
    if pd.isna(raw_value) or raw_value is None:
        return ""
    try:
        return str(int(float(raw_value))).zfill(2)
    except (ValueError, TypeError):
        return str(raw_value).strip()


class RegistryMismatchError(ValueError):
    """Raised when the uploaded file doesn't look like a Credit registry."""
    pass


# CustomerNumber is required because the loop below skips (via
# `if not customer_number: continue`) any row where it's missing — so a
# file without this column would previously produce a "successful" run
# with zero rows written, instead of an error.
REQUIRED_CREDIT_COLUMNS = ['CustomerNumber', 'Risk Cat', 'Credit rep.group', 'Credit Limit']


def validate_credit_registry(df):
    missing = [c for c in REQUIRED_CREDIT_COLUMNS if c not in df.columns]
    if missing:
        raise RegistryMismatchError(
            "This file doesn't look like a Credit registry. "
            f"Missing expected column(s): {', '.join(missing)}. "
            "Make sure you selected the right file for the Credit Management process."
        )


def process_credit_registry(
    registry_file,
    template_path="templates/credit_load_template.xlsx",
    but_path="reference_data/but0id_qs4_500.xlsx"
) -> io.BytesIO:
    """
    Processes the Credit Registry spreadsheet and populates:

        1. Profile - Credit MD for Cust.
        2. Segment - Credit MD for Cust.

    Registry -> Credit Profile mapping:

        KUNNR              -> CustomerNumber -> BUT.Business Partner
        RUN_ID             -> Sequential
        OWN_RATING         -> Blank
        CHECK_RULE         -> "01"
        LIMIT_RULE         -> "SAP_ALL"
        RATING_VAL_DATE    -> Blank
        RISK_CLASS         -> Risk Cat
        CREDIT_GROUP       -> Credit rep.group (read directly from the
                              registry — the DAP Clerk Codes lookup file
                              is no longer used; the registry now carries
                              this value itself)

    Registry -> Credit Segment mapping:

        KUNNR              -> CustomerNumber -> BUT.Business Partner
        RUN_ID             -> Sequential
        CREDIT_SEGMENT     -> "1000"
        CREDIT_LIMIT       -> Credit Limit

    Data is written starting from row 9.
    """

    # ---------------------------------------------------------
    # 1. Read Registry
    # ---------------------------------------------------------

    df_raw = pd.read_excel(registry_file)

    # Fail loudly if this isn't actually a Credit registry, instead of
    # silently skipping every row and returning an empty "success".
    validate_credit_registry(df_raw)

    # Customer Number -> Business Partner, via the BUT reference sheet,
    # scoped to the 'DAP' (customer) Identification Type — see
    # reference_mappings.load_but_mapping() for why the id_type scoping
    # matters (the same Identification Number can resolve to a different
    # Business Partner under 'DAPVEN').
    customer_but_mapping = load_but_mapping(
        but_path,
        id_type="DAP"
    )

    # ---------------------------------------------------------
    # 2. Load Credit Data Load template
    # ---------------------------------------------------------

    wb = openpyxl.load_workbook(template_path)

    sheets_to_fill = [
        "Profile - Credit MD for Cust.",
        "Segment - Credit MD for Cust."
    ]

    # ---------------------------------------------------------
    # 3. Prepare mapped records
    # ---------------------------------------------------------

    processed_records = []

    run_id = 1

    for idx, row in df_raw.iterrows():

        raw_customer_number = clean_string(
            row.get("CustomerNumber")
        )

        # Skip rows where CustomerNumber is blank
        if not raw_customer_number:
            continue

        record = {
            # KUNNR in the output is the S/4 Business Partner, resolved
            # from the registry's raw Customer Number via the BUT sheet
            # — not the raw registry value itself.
            # "customer_number": map_business_partner(
            #     customer_but_mapping,
            #     raw_customer_number
            # ),
            "customer_number": raw_customer_number,

            # Kept for traceability (validation messages, source lookups)
            # — this is the key the BUT and Clerk Codes lookups both use.
            "source_customer_number": raw_customer_number,

            "run_id": str(run_id),

            # +2: pandas index is 0-based and Row 1 is the header, so the
            # first data row (index 0) is Excel Row 2.
            "source_row": idx + 2,

            # Profile fields
            "risk_class": clean_string(
                row.get("Risk Cat")
            ),

            # Credit Rep Group is read directly from the registry's own
            # 'Credit rep.group' column — the DAP Clerk Codes lookup file
            # is no longer used for this field.
            "credit_group": normalize_credit_rep_group(
                row.get("New Credit Rep Group")
            ),

            # Segment fields
            "credit_limit": clean_float(
                row.get("Credit Limit")
            )
        }

        processed_records.append(record)

        run_id += 1

    # ---------------------------------------------------------
    # 4. Populate both sheets
    # ---------------------------------------------------------

    validation_errors = []

    for sheet_name in sheets_to_fill:

        if sheet_name not in wb.sheetnames:
            print(
                f"Sheet {sheet_name} not found in template, skipping."
            )
            continue

        ws = wb[sheet_name]

        # -----------------------------------------------------
        # Technical field names are on Row 5
        # -----------------------------------------------------

        tech_cols = [
            clean_string(
                ws.cell(row=5, column=col).value
            )
            for col in range(1, ws.max_column + 1)
        ]

        # Build:
        # technical_field_name -> Excel column number
        col_to_idx = {
            name: idx + 1
            for idx, name in enumerate(tech_cols)
            if name
        }

        # Which technical fields this sheet marks mandatory (Row 8, '*').
        mandatory_fields = extract_mandatory_fields(ws, col_to_idx)

        # -----------------------------------------------------
        # Data starts from Row 9
        # -----------------------------------------------------

        current_row = 9

        for rec in processed_records:

            mapped_values = {}

            # =================================================
            # PROFILE SHEET
            # =================================================

            if sheet_name == "Profile - Credit MD for Cust.":

                mapped_values["KUNNR"] = rec["customer_number"]

                mapped_values["RUN_ID"] = rec["run_id"]

                # No mapping provided for OWN_RATING
                mapped_values["OWN_RATING"] = ""

                # Hardcoded for now
                mapped_values["CHECK_RULE"] = "02"

                # Hardcoded
                mapped_values["LIMIT_RULE"] = "CRED_SCORE"

                # Blank as specified
                mapped_values["RATING_VAL_DATE"] = ""

                # Risk Cat from Registry
                mapped_values["RISK_CLASS"] = rec["risk_class"]

                # Credit rep.group must remain STRING
                mapped_values["CREDIT_GROUP"] = rec["credit_group"]

            # =================================================
            # SEGMENT SHEET
            # =================================================

            elif sheet_name == "Segment - Credit MD for Cust.":

                mapped_values["KUNNR"] = rec["customer_number"]

                mapped_values["RUN_ID"] = rec["run_id"]

                # Hardcoded for now
                mapped_values["CREDIT_SEGMENT"] = "1000"

                mapped_values["CREDIT_LIMIT"] = rec["credit_limit"]

                # Other fields are intentionally left blank
                # because no mapping was provided.

            # -------------------------------------------------
            # Flag any mandatory field that's blank in the final mapped
            # value. The sheet still gets written either way; this only
            # records the issue.
            # -------------------------------------------------

            for field_tech, field_label in mandatory_fields.items():
                if is_blank(mapped_values.get(field_tech)):
                    validation_errors.append({
                        'sheet': sheet_name,
                        'field': field_tech,
                        'field_label': field_label,
                        'source_row': rec['source_row'],
                        'customer_number': rec['customer_number'],
                        'source_customer_number': rec['source_customer_number'],
                    })

            # -------------------------------------------------
            # Write mapped values
            # -------------------------------------------------

            for col_name, value in mapped_values.items():

                if col_name in col_to_idx:

                    c_idx = col_to_idx[col_name]

                    ws.cell(
                        row=current_row,
                        column=c_idx,
                        value=value
                    )

            current_row += 1

    # ---------------------------------------------------------
    # 5. Save workbook to BytesIO
    # ---------------------------------------------------------

    out_buf = io.BytesIO()

    wb.save(out_buf)

    out_buf.seek(0)

    return out_buf, validation_errors