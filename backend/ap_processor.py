import io
import datetime

import pandas as pd
import openpyxl

import mappings
from validation_utils import extract_mandatory_fields, is_blank
from reference_mappings import load_but_mapping, map_business_partner


# ============================================================
# Common Cleaning Functions
# ============================================================

def clean_string(val):
    """
    Convert a value to a clean string.

    Prevents values such as 5304994.0 from being written
    when the Excel source contains an integer-like float.
    """
    if pd.isna(val) or val is None:
        return ""

    if isinstance(val, float) and val.is_integer():
        return str(int(val))

    return str(val).strip()


def clean_int(val, default=""):
    """
    Convert a value to an integer.

    Blank / NaN values return the supplied default.
    """
    if pd.isna(val) or val is None:
        return default

    try:
        return int(float(val))
    except (ValueError, TypeError):
        return str(val).strip()


def clean_float(val, default=None):
    """
    Convert a value to float, including SAP/Excel "accounting format"
    text such as "5,114.43-" or "(5,114.43)" for negative numbers.

    Registry exports commonly render negatives with a trailing minus
    sign and/or thousands separators rather than a leading minus.
    Python's float() can't parse either of those directly, so without
    this such values fall through unchanged as literal text (with the
    trailing minus baked in) instead of becoming the real negative
    number -5114.43 — which matters here since apply_debit_credit_sign()
    below relies on being handed an already-parsed float.
    """
    if pd.isna(val) or val is None:
        return default

    if isinstance(val, (int, float)):
        return float(val)

    text = str(val).strip()
    if not text:
        return default

    negative = False

    if text.endswith('-'):
        negative = True
        text = text[:-1].strip()
    elif text.startswith('(') and text.endswith(')'):
        negative = True
        text = text[1:-1].strip()

    text = text.replace(',', '')

    try:
        num = float(text)
        return -num if negative else num
    except (ValueError, TypeError):
        return val


def clean_date(val):
    """
    Convert Excel/date values into Python date objects.
    """

    if pd.isna(val) or val is None:
        return None

    if isinstance(val, (datetime.date, datetime.datetime)):
        return (
            val.date()
            if isinstance(val, datetime.datetime)
            else val
        )

    val_str = str(val).strip()

    if val_str in (
        "",
        "00/00/0000",
        "00.00.0000",
        "NaT"
    ):
        return None

    date_formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%d.%m.%Y",
        "%Y%m%d",
    )

    for fmt in date_formats:
        try:
            return datetime.datetime.strptime(
                val_str,
                fmt
            ).date()
        except ValueError:
            continue

    return val


# ============================================================
# AP Payment Terms Mapping
# ECC -> S/4
# ============================================================

PAYMENT_TERMS_MAPPING = {
    "O": "NT30",
    "A": "NT00",
    "B": "NT10",
    "H": "NT15",
    "T": "NT60",
    "C": "Z130",
    "L": "NT20",
    "YY": "NT90",
    "R": "NT45",
    "S": "NT50",
    "NF5": "Z514",
    "D": "Z221",
    "ZZ": "N100",
    "I": "Z229",
    "N120": "N120",
    "M": "Z120",
    "EE": "Z167",
    "BB": "Z103",
    "G": "Z162",
    "NF7": "NT07",
    "HI": "Z101",
    "U": "Z053",
    "N110": "N110",
    "TT": "NT75",
    "Y": "NT55",
    "V": "Z233",
    "Q": "NT40",
    "X": "Z247",
    "WX": "Z163",
    "N115": "N115",
    "Z": "Z132",
    "J": "P215",
    "E10": "P210",
    "E": "P010",
    "FF": "Z145",
    "N125": "N125",
    "N65": "NT65",
    "T70": "NT70",
    "NF4": "Z333",
    "N135": "N135",
    "AA": "Z261",
    "CC": "Z147",
    "W": "Z263",
    "DD": "Z100",
    "F": "P231",
    "XX": "Z190",
    "N": "Z223",
    "OO": "NT38",
}


def get_s4_payment_terms(ecc_payment_term):
    """
    Look up S/4 Payment Terms from ECC Payment Terms.
    """

    if not ecc_payment_term or pd.isna(ecc_payment_term):
        return ""

    payment_term = str(ecc_payment_term).strip().upper()

    return PAYMENT_TERMS_MAPPING.get(
        payment_term,
        payment_term
    )


# ============================================================
# Company Code Mapping
# ECC -> S/4
# ============================================================

def get_s4_company_code(ecc_company_code):
    """
    Convert ECC Company Code to S/4 Company Code
    using the existing mappings.py logic.
    """

    if not ecc_company_code or pd.isna(ecc_company_code):
        return ""

    return clean_string(
        mappings.get_s4_company_code(ecc_company_code)
    )


# ============================================================
# Tax Code Mapping
# ============================================================

def get_s4_tax_code(ecc_company_code, s4_company_code):
    """
    Tax Code is determined based on the Company Code.

    ECC CA01 -> S/4 1200 -> Z0
    ECC US01 -> S/4 1000 -> I0
    ECC US06 -> S/4 1001 -> I0
    """

    ecc_company_code = clean_string(
        ecc_company_code
    ).upper()

    s4_company_code = clean_string(
        s4_company_code
    )

    # ECC -> S/4 based tax-code rules
    if ecc_company_code == "CA01" or s4_company_code == "1200":
        return "Z0"

    if ecc_company_code in ("US01", "US06") or s4_company_code in (
        "1000",
        "1001"
    ):
        return "I0"

    return ""


# ============================================================
# Debit/Credit Indicator (SHKZG) Sign Logic
# ============================================================

def apply_debit_credit_sign(amount, indicator):
    """
    Apply the sign implied by the Debit/Credit Indicator (SHKZG) to an
    amount value.

        SHKZG == "S" (Debit)  -> amount is positive
        SHKZG == "H" (Credit) -> amount is negative

    Magnitude is always preserved; only the sign is normalized. The
    indicator is treated as the authoritative source of truth for sign,
    so this overrides whatever sign the source amount already had rather
    than trusting it. Any indicator other than "S"/"H" is returned as a
    plain float (magnitude and sign both left alone) instead of guessing.

    A positive result is a plain Python float (e.g. 500.0, not "+500.0")
    — Python never prefixes a positive number with "+", so no extra
    handling is needed to keep the sign off positive values.
    """
    if amount is None:
        return amount

    try:
        magnitude = abs(float(amount))
    except (TypeError, ValueError):
        return amount

    ind = clean_string(indicator).upper()

    if ind == "H":
        return -magnitude
    if ind == "S":
        return magnitude

    return amount


# ============================================================
# Registry Shape Validation
# ============================================================

class RegistryMismatchError(ValueError):
    """Raised when the uploaded file doesn't look like an AP registry."""
    pass


REQUIRED_AP_COLUMNS = [
    "Company Code",
    "Supplier",
    "Amount",
    "Reference"
]

# Stable SAP technical field -> canonical field name used by the processor
AP_TECHNICAL_TO_CANONICAL = {
    "BUKRS": "Company Code",
    "LIFNR": "Supplier",
    "BUZEI": "Line Item", ### New added by me
    "BLDAT": "Document Date",
    "WAERS": "Currency",
    "XBLNR": "Reference",
    "SGTXT": "Text",
    "DMBTR": "Amt.in loc.cur.",
    "WRBTR": "Amount",
    "MWSKZ": "Tax Code",
    "ZTERM": "Terms of Payment",
    "ZFBDT": "Baseline Payment Dte",
    "ZLSCH": "Payment Method",
    "ZLSPR": "Payment Block",
    "ZBD1T": "Days 1",
    "ZBD1P": "Disc.percent 1",
    "ZBD2T": "Days 2",
    "ZBD2P": "Disc.percent 2",
    "ZBD3T": "Days Net",
    "SKFBT": "Discount base",
    "DTWS1": "Instruction 1",
    "DTWS2": "Instruction 2",
    "DTWS3": "Instruction 3",
    "DTWS4": "Instruction 4",
    "XREF1": "Reference Key 1",
    "BELNR": "Document Number",
    "BLART": "Document Type",
}


def read_ap_registry(registry_file):
    """
    Read an AP registry regardless of whether the export contains:

    Format 1:
        Row 1 = technical SAP field names
        Row 2 = human-readable names
        Row 3+ = data

    Format 2:
        Row 1 = technical SAP field names
        Row 2+ = data

    The returned DataFrame always uses the canonical human-readable
    column names expected by the processor.
    """

    # Read the first two rows without assuming a header
    preview = pd.read_excel(
        registry_file,
        header=None,
        nrows=2
    )

    registry_file.seek(0)

    if preview.empty:
        raise RegistryMismatchError(
            "The uploaded AP registry is empty."
        )

    # ---------------------------------------------------------
    # Check Row 1 for technical SAP field names
    # ---------------------------------------------------------

    row0 = [
        str(v).strip()
        for v in preview.iloc[0]
        if pd.notna(v)
    ]

    technical_matches = sum(
        1
        for col in AP_TECHNICAL_TO_CANONICAL
        if col in row0
    )

    # ---------------------------------------------------------
    # CASE 1:
    # Row 1 contains technical SAP fields
    # ---------------------------------------------------------

    if technical_matches >= 3:

        # Read Row 1 as the actual header
        df = pd.read_excel(
            registry_file,
            header=0
        )

        # Check whether Row 2 is actually the human-readable
        # description row rather than a real data row.
        if len(preview) > 1:

            row1 = [
                str(v).strip()
                for v in preview.iloc[1]
                if pd.notna(v)
            ]

            human_matches = sum(
                1
                for col in REQUIRED_AP_COLUMNS
                if col in row1
            )

            # If Row 2 contains labels such as
            # "Company Code", "Supplier", "Amount", "Reference",
            # remove it because pandas currently considers it data.
            if human_matches >= 2:
                df = df.iloc[1:].copy()

        # Rename the stable technical SAP fields to the canonical
        # names expected by the rest of the processor.
        df = df.rename(
            columns=AP_TECHNICAL_TO_CANONICAL
        )

    # ---------------------------------------------------------
    # CASE 2:
    # Row 1 already contains human-readable fields
    # ---------------------------------------------------------

    else:

        row0_values = set(row0)

        row1_values = (
            {
                str(v).strip()
                for v in preview.iloc[1]
                if pd.notna(v)
            }
            if len(preview) > 1
            else set()
        )

        row0_matches = sum(
            1
            for col in REQUIRED_AP_COLUMNS
            if col in row0_values
        )

        row1_matches = sum(
            1
            for col in REQUIRED_AP_COLUMNS
            if col in row1_values
        )

        if row0_matches == 0 and row1_matches == 0:
            raise RegistryMismatchError(
                "This file doesn't look like an AP registry. "
                "The expected AP fields could not be identified."
            )

        # Human-readable headers are on Row 2
        if row1_matches > row0_matches:
            df = pd.read_excel(
                registry_file,
                header=1
            )

        # Human-readable headers are on Row 1
        else:
            df = pd.read_excel(
                registry_file,
                header=0
            )

    # ---------------------------------------------------------
    # Final validation
    # ---------------------------------------------------------

    missing = [
        col
        for col in REQUIRED_AP_COLUMNS
        if col not in df.columns
    ]

    if missing:
        raise RegistryMismatchError(
            "This file looks like an AP registry, but some required "
            f"fields are missing: {', '.join(missing)}."
        )

    return df


# ============================================================
# Main AP Processor
# ============================================================

def process_ap_registry(
    registry_file,
    template_path="templates/AP Data Load Sheet - SIT2.xlsx",
    but_path="reference_data/but0id_qs4_500.xlsx"
) -> io.BytesIO:
    """
    Processes the AP Registry Excel file and populates:

        1. Vendor Open Items
        2. Withholding Tax Items

    ============================================================
    VENDOR OPEN ITEMS
    ============================================================

    BUKRS       <- Company Code -> ECC to S/4 mapping
    XBLNR       <- Reference
    DOCLN       <- Blank
    LIFNR       <- Supplier
    GKONT       <- 9999900000
    BLART       <- UE
    BLDAT       <- Document Date
    SGTXT       <- Text
    WAERS       <- Currency
    WRBTR       <- Amount
    DMBTR       <- Amt.in loc.cur.
    DMBE2       <- LC2 Amount
    DMBE3       <- LC3 Amount
    MWSKZ       <- Company Code based tax mapping
    ZTERM       <- Terms of Payment -> ECC to S/4 mapping
    ZFBDT       <- Baseline Payment Dte
    ZLSCH       <- Payment Method
    ZLSPR       <- Payment Block
    ZBD1T       <- Days 1
    ZBD1P       <- Disc.percent 1
    ZBD2T       <- Days 2
    ZBD2P       <- Disc.percent 2
    ZBD3T       <- Days Net
    SKFBT       <- Discount base
    DTWS1       <- Instruction 1
    DTWS2       <- Instruction 2
    DTWS3       <- Instruction 3
    DTWS4       <- Instruction 4
    XREF1       <- Reference Key 1

    All other Vendor Open Items fields remain blank.


    ============================================================
    WITHHOLDING TAX ITEMS
    ============================================================

    BUKRS       <- Company Code -> ECC to S/4 mapping
    XBLNR       <- Reference
    DOCLN       <- Blank
    LIFNR       <- Supplier
    WT_TYPE     <- Blank
    WT_CODE     <- Blank
    BAS_AMT_TC  <- Blank
    MAN_AMT_TC  <- Blank

    All other Withholding Tax Items fields remain blank.

    Data is written starting from Row 9.
    Technical field identifiers are read from Row 5.
    """

    # ---------------------------------------------------------
    # 1. Read AP Registry
    # ---------------------------------------------------------

    # Detects whether the human-readable labels are on Row 1 or Row 2 of
    # this particular export and reads with the correct header row — see
    # read_ap_registry() docstring for why this matters.
    df_raw = read_ap_registry(
        registry_file
    )

    # ---------------------------------------------------------
    # 2. Remove accidental/header row if present
    #
    # The uploaded AP Registry contains a first data row
    # containing technical field names such as:
    #
    # Company Code = BUKRS
    # Supplier = LIFNR
    # Reference = XBLNR
    #
    # This must not become an actual migration record. Only relevant when
    # the technical-code row landed on Row 1 (i.e. it became Row 2 of
    # the data) — if it was already excluded as the header itself
    # (Row 2-as-header case), there's nothing to strip here.
    # ---------------------------------------------------------

    if "Company Code" in df_raw.columns:

        df_raw = df_raw[
            df_raw["Company Code"]
            .astype(str)
            .str.strip()
            .str.upper()
            != "BUKRS"
        ].copy()

    # ---------------------------------------------------------
    # 3. Load AP Data Load template
    # ---------------------------------------------------------

    wb = openpyxl.load_workbook(
        template_path
    )

    # Supplier -> Business Partner, via the BUT reference sheet, scoped
    # to the 'DAPVEN' (vendor) Identification Type. Built once and reused
    # for both Vendor Open Items and Withholding Tax Items below — see
    # reference_mappings.load_but_mapping() for why the id_type scoping
    # matters.
    vendor_but_mapping = load_but_mapping(
        but_path,
        id_type="DAP"
    )

    # =========================================================
    # Helper function to prepare technical columns
    # =========================================================

    def get_column_mapping(ws):

        tech_cols = [
            clean_string(
                ws.cell(
                    row=5,
                    column=col
                ).value
            )
            for col in range(
                1,
                ws.max_column + 1
            )
        ]

        return {
            name: idx + 1
            for idx, name in enumerate(tech_cols)
            if name
        }

    # =========================================================
    # Helper function to clear existing template data
    # =========================================================

    def clear_data_rows(ws):

        # Existing template contains example/sample rows.
        # Clear their values while preserving formatting.

        for row in range(
            9,
            ws.max_row + 1
        ):
            for col in range(
                1,
                ws.max_column + 1
            ):
                ws.cell(
                    row=row,
                    column=col
                ).value = None

    # =========================================================
    # 4. Vendor Open Items
    # =========================================================

    validation_errors = []

    vendor_sheet = "Vendor Open Items"

    if vendor_sheet not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{vendor_sheet}' not found in AP template."
        )

    ws_vendor = wb[vendor_sheet]

    vendor_col_to_idx = get_column_mapping(
        ws_vendor
    )

    # Which technical fields this sheet marks mandatory (Row 8, '*').
    vendor_mandatory_fields = extract_mandatory_fields(
        ws_vendor, vendor_col_to_idx
    )

    # Clear sample data from template
    clear_data_rows(
        ws_vendor
    )

    current_row = 9

    for idx, row_data in df_raw.iterrows():

        # -----------------------------------------------------
        # ECC Company Code -> S/4 Company Code
        # -----------------------------------------------------

        ecc_company_code = clean_string(
            row_data.get("Company Code")
        )

        s4_company_code = get_s4_company_code(
            ecc_company_code
        )

        xblnr_value = clean_string(row_data.get("Reference"))
        if not xblnr_value:
            xblnr_value = "No reference in ECC"

        # -----------------------------------------------------
        # Tax Code based on Company Code
        # -----------------------------------------------------

        s4_tax_code = get_s4_tax_code(
            ecc_company_code,
            s4_company_code
        )

        # -----------------------------------------------------
        # Payment Terms ECC -> S/4
        # -----------------------------------------------------

        s4_payment_terms = get_s4_payment_terms(
            row_data.get("Terms of Payment")
        )

        # -----------------------------------------------------
        # Build mapped values
        #
        # Everything not listed remains blank.
        # -----------------------------------------------------

        # -----------------------------------------------------
        # Reference Key 1 logic
        # -----------------------------------------------------

        document_type = clean_string(
            row_data.get("Document Type")
        ).upper()

        if document_type in ("KR", "RE"):
            xref1_value = clean_string(
                row_data.get("Document Number")
            )
        else:
            xref1_value = clean_string(
                row_data.get("Reference Key 1")
            )

        # -----------------------------------------------------
        # Debit/Credit Indicator sign logic (SHKZG)
        #
        # S -> amount positive, H -> amount negative. Applied to both
        # Amount (WRBTR) and Amt.in loc.cur. (DMBTR) — the indicator
        # decides the sign regardless of what sign the source value
        # already carried.
        # -----------------------------------------------------

        shkzg = row_data.get("SHKZG")

        s4_amount = apply_debit_credit_sign(
            clean_float(row_data.get("Amount")),
            shkzg
        )

        s4_amount_lc = apply_debit_credit_sign(
            clean_float(row_data.get("Amt.in loc.cur.")),
            shkzg
        )

        # -----------------------------------------------------
        # Build mapped values
        #
        # Everything not listed remains blank.
        # -----------------------------------------------------


        mapped_values = {

            # Company Code
            "BUKRS": s4_company_code,

            # Reference
            # "XBLNR": clean_string(
            #     row_data.get("Reference")
            # ),
            "XBLNR": xblnr_value,

            # Line Item Number
            "DOCLN": clean_string(
                row_data.get("Line Item")
            ),

            # Supplier -> Business Partner (via BUT reference sheet)
            # "LIFNR": map_business_partner(
            #     vendor_but_mapping,
            #     row_data.get("Supplier")
            # ),
            "LIFNR": clean_string(
                row_data.get("Supplier")
            ),

            # GKONT - Hardcoded
            "GKONT": "9999900000",

            # BLART - Hardcoded
            "BLART": "UE",

            # Document Date
            "BLDAT": clean_date(
                row_data.get("Document Date")
            ),

            # Text
            "SGTXT": clean_string(
                row_data.get("Text")
            ),

            # Currency
            "WAERS": clean_string(
                row_data.get("Currency")
            ),

            # Amount - signed based on Debit/Credit Indicator (SHKZG):
            # S -> positive, H -> negative
            "WRBTR": s4_amount,

            "HWAER": clean_string(
                row_data.get("Currency")
            ),

            # Local Currency Amount - same SHKZG-based sign as WRBTR
            "DMBTR": s4_amount_lc,

            # LC2 Amount
            "DMBE2": clean_float(
                row_data.get("LC2 Amount")
            ),

            # LC3 Amount
            "DMBE3": clean_float(
                row_data.get("LC3 Amount")
            ),

            # Tax Code - based on company code
            "MWSKZ": s4_tax_code,

            # Payment Terms - ECC -> S/4
            "ZTERM": s4_payment_terms,

            # Baseline Payment Date
            "ZFBDT": clean_date(
                row_data.get("Baseline Payment Dte")
            ),

            # Payment Method
            "ZLSCH": clean_string(
                row_data.get("Payment Method")
            ),

            # Payment Block
            "ZLSPR": clean_string(
                row_data.get("Payment Block")
            ),

            # Cash Discount Days 1
            "ZBD1T": clean_int(
                row_data.get("Days 1")
            ),

            # Cash Discount Percentage 1
            "ZBD1P": clean_float(
                row_data.get("Disc.percent 1")
            ),

            # Cash Discount Days 2
            "ZBD2T": clean_int(
                row_data.get("Days 2")
            ),

            # Cash Discount Percentage 2
            "ZBD2P": clean_float(
                row_data.get("Disc.percent 2")
            ),

            # Net Due Days
            "ZBD3T": clean_int(
                row_data.get("Days Net")
            ),

            # Discount Base
            "SKFBT": clean_float(
                row_data.get("Discount base")
            ),

            # Instruction 1
            "DTWS1": clean_string(
                row_data.get("Instruction 1")
            ),

            # Instruction 2
            "DTWS2": clean_string(
                row_data.get("Instruction 2")
            ),

            # Instruction 3
            "DTWS3": clean_string(
                row_data.get("Instruction 3")
            ),

            # Instruction 4
            "DTWS4": clean_string(
                row_data.get("Instruction 4")
            ),

            # Reference Key 1
            # "XREF1": clean_string(
            #     row_data.get("Reference Key 1")
            # ),
            # Reference Key 1
            "XREF1": xref1_value,
        }

        # -----------------------------------------------------
        # Flag any mandatory field that's blank in the final mapped
        # value. The sheet still gets written either way; this only
        # records the issue.
        # -----------------------------------------------------

        for field_tech, field_label in vendor_mandatory_fields.items():
            if is_blank(mapped_values.get(field_tech)):
                validation_errors.append({
                    'sheet': vendor_sheet,
                    'field': field_tech,
                    'field_label': field_label,
                    # +2: pandas index is 0-based and Row 1 of the
                    # registry is the header.
                    'source_row': idx + 2,
                    'company_code': mapped_values.get('BUKRS'),
                    'vendor': mapped_values.get('LIFNR'),
                    'reference': mapped_values.get('XBLNR'),
                })

        # -----------------------------------------------------
        # Write values to Vendor Open Items
        # -----------------------------------------------------

        for col_name, value in mapped_values.items():

            if col_name in vendor_col_to_idx:

                ws_vendor.cell(
                    row=current_row,
                    column=vendor_col_to_idx[col_name],
                    value=value
                )

        current_row += 1

    # =========================================================
    # 5. Withholding Tax Items
    # =========================================================

    withholding_sheet = "Withholding Tax Items"

    if withholding_sheet not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{withholding_sheet}' not found in AP template."
        )

    ws_withholding = wb[withholding_sheet]

    withholding_col_to_idx = get_column_mapping(
        ws_withholding
    )

    # Which technical fields this sheet marks mandatory (Row 8, '*').
    withholding_mandatory_fields = extract_mandatory_fields(
        ws_withholding, withholding_col_to_idx
    )

    # Clear sample data from template
    clear_data_rows(
        ws_withholding
    )

    current_row = 9

    for idx, row_data in df_raw.iterrows():

        # -----------------------------------------------------
        # ECC Company Code -> S/4 Company Code
        # -----------------------------------------------------

        ecc_company_code = clean_string(
            row_data.get("Company Code")
        )

        s4_company_code = get_s4_company_code(
            ecc_company_code
        )

        # Compute XBLNR with fallback (repeat)
        xblnr_value = clean_string(row_data.get("Reference"))
        if not xblnr_value:
            xblnr_value = "No reference in ECC"

        # -----------------------------------------------------
        # Withholding Tax mapping
        #
        # Only the explicitly requested fields are populated.
        # -----------------------------------------------------

        mapped_values = {

            # Company Code
            "BUKRS": s4_company_code,

            # Reference
            # "XBLNR": clean_string(
            #     row_data.get("Reference")
            # ),
            "XBLNR": xblnr_value,

            # Explicitly blank
            # Line Item Number
            "DOCLN": clean_string(
                row_data.get("Line Item")
            ),

            # Supplier -> Business Partner (via BUT reference sheet)
            # "LIFNR": map_business_partner(
            #     vendor_but_mapping,
            #     row_data.get("Supplier")
            # ),
            "LIFNR": clean_string(
                row_data.get("Supplier")
            ),

            # Explicitly blank
            "WT_TYPE": "",
            "WT_CODE": "",
            "BAS_AMT_TC": "",
            "MAN_AMT_TC": "",
        }

        # -----------------------------------------------------
        # Flag any mandatory field that's blank in the final mapped
        # value. The sheet still gets written either way; this only
        # records the issue.
        # -----------------------------------------------------

        for field_tech, field_label in withholding_mandatory_fields.items():
            if is_blank(mapped_values.get(field_tech)):
                validation_errors.append({
                    'sheet': withholding_sheet,
                    'field': field_tech,
                    'field_label': field_label,
                    'source_row': idx + 2,
                    'company_code': mapped_values.get('BUKRS'),
                    'vendor': mapped_values.get('LIFNR'),
                    'reference': mapped_values.get('XBLNR'),
                })

        # -----------------------------------------------------
        # Write values to Withholding Tax Items
        # -----------------------------------------------------

        for col_name, value in mapped_values.items():

            if col_name in withholding_col_to_idx:

                ws_withholding.cell(
                    row=current_row,
                    column=withholding_col_to_idx[col_name],
                    value=value
                )

        current_row += 1

    # =========================================================
    # 6. Save workbook
    # =========================================================

    out_buf = io.BytesIO()

    wb.save(
        out_buf
    )

    out_buf.seek(0)

    return out_buf, validation_errors